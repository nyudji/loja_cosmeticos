import streamlit as st
import pandas as pd
from datetime import datetime, date
import os
import uuid
import calendar
from pandas.tseries.offsets import DateOffset 
import plotly.express as px
import numpy as np
import time
import openpyxl
from openpyxl.utils import get_column_letter


# ===== CONSTANTES DO main.py (Preservadas e Modificadas) =====
# Caminho do arquivo Excel
ARQUIVO_EXCEL = os.path.join("app", "dados", "BD_Loja.xlsx")
PRODUTOS_EXCEL = os.path.join("app", "dados", "BD_Loja.xlsx")

# Colunas do Movimento: Data Pagamento Removida
COLUNAS = [
    "Data", "COD do Produto", "Produto", "Cliente", "Tipo de Movimentação",
    "Quantidade", "Preço Custo Total", "Preço Venda Total", "Observações", "Status",
    "Data Prevista", "Tipo de Pagamento", "ID_Venda" 
]
# =============================================================


# ===== FUNÇÕES DE I/O DO main.py (Preservadas e Modificadas) =====

def carregar_dados():
    """Carrega os dados da planilha Excel, cria se não existir ou estiver corrompido."""
    os.makedirs(os.path.dirname(ARQUIVO_EXCEL), exist_ok=True)
    
    dtype_force = {
        "COD do Produto": str,
        "Produto": str,
        "Cliente": str,
        "Observações": str,
        "Data Prevista": str, 
        "Preço Custo Total": float,
        "Preço Venda Total": float
    }

    if not os.path.exists(ARQUIVO_EXCEL):
        df_vazio = pd.DataFrame(columns=COLUNAS)
        with pd.ExcelWriter(ARQUIVO_EXCEL, engine='openpyxl') as writer:
            df_vazio.to_excel(writer, sheet_name="Movimento", index=False)
        return df_vazio

    try:
        xls = pd.ExcelFile(ARQUIVO_EXCEL)
        if "Movimento" not in xls.sheet_names:
            raise ValueError("Aba 'Movimento' não encontrada no arquivo Excel.")
            
        df = pd.read_excel(xls, sheet_name="Movimento", dtype=dtype_force)
        
        df['Data'] = pd.to_datetime(df['Data'], errors='coerce')
        # Tenta converter a única coluna de data para datetime
        df['Data Prevista'] = pd.to_datetime(df['Data Prevista'], errors='coerce') 
        
        df_final = df[[col for col in COLUNAS if col in df.columns]]
        for col in COLUNAS:
            if col not in df_final.columns:
                df_final[col] = None
        
        return df_final[COLUNAS] 

    except Exception as e:
        st.error(f"Erro ao carregar dados: {e}. Criando arquivo novo.")
        df_vazio = pd.DataFrame(columns=COLUNAS)
        with pd.ExcelWriter(ARQUIVO_EXCEL, engine='openpyxl') as writer:
            df_vazio.to_excel(writer, sheet_name="Movimento", index=False)
        return df_vazio

# DEFINIÇÃO CORRIGIDA DE salvar_dados: Retorna True/False
def salvar_dados(df):
    """Salva o DataFrame de volta na planilha Excel. Retorna True ou False."""
    try:
        df_salvar = df.copy()
        for col in COLUNAS:
             if col not in df_salvar.columns:
                 df_salvar[col] = None

        with pd.ExcelWriter(ARQUIVO_EXCEL, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df_salvar[COLUNAS].to_excel(writer, sheet_name='Movimento', index=False)
        st.success("Dados salvos com sucesso!")
        return True 
    except Exception as e:
        st.error(f"Erro ao salvar dados: {e}")
        return False 
    
# *** A DEFINIÇÃO DUPLICADA ABAIXO FOI REMOVIDA ***
# def salvar_dados(df):
#     """Salva o DataFrame de volta na planilha Excel."""
#     try:
#         ...
#         st.success("Dados salvos com sucesso!")
#     except Exception as e:
#         st.error(f"Erro ao salvar dados: {e}")
# *************************************************

def carregar_produtos():
    """Carrega a base de produtos (mantida do main.py para compatibilidade)."""
    try:
        xls = pd.ExcelFile(PRODUTOS_EXCEL)
        if "Produtos" not in xls.sheet_names:
            st.error("Aba 'Produtos' não encontrada no arquivo Excel.")
            return pd.DataFrame()
        
        # Carrega todas as colunas da aba Produtos
        df_produtos = pd.read_excel(xls, sheet_name="Produtos").fillna('')
        
        # Definição das colunas usadas para construir o nome do produto (se não houver um)
        COL_D = 'Categoria' 
        COL_C = 'Coleção' 
        COL_F = 'Nome' 
        COL_G = 'Unidade' 
        COL_H = 'Volume' 
        COL_I = 'Tipo' 
        
        # Verifica se 'Produto' existe e usa como base
        df_produtos['Produto_Final'] = df_produtos['Produto'].astype(str).str.strip() if 'Produto' in df_produtos.columns else ''
        
        colunas_formula = [COL_D, COL_C, COL_F, COL_G, COL_H, COL_I]
        if all(col in df_produtos.columns for col in colunas_formula):
            
            def recalcular_produto(row):
                if row['Produto_Final'] == '' or row['Produto_Final'].lower() == 'nan':
                    parte_gh = str(row[COL_H]).strip()
                    try:
                        val_g = float(str(row[COL_G]).replace(',', '.').strip())
                        if val_g > 0:
                            parte_gh = f"{val_g}x{row[COL_H].strip()}"
                    except:
                        pass
                    
                    partes = [
                        str(row[COL_D]).strip(),
                        str(row[COL_C]).strip(),
                        str(row[COL_F]).strip(),
                        parte_gh,
                        str(row[COL_I]).strip()
                    ]
                    nome_recalculado = ' '.join(filter(None, partes))
                    return nome_recalculado if nome_recalculado != '' else 'Produto sem nome na planilha'
                else:
                    return row['Produto_Final']

            df_produtos['Produto_Final'] = df_produtos.apply(recalcular_produto, axis=1)
        
        df_produtos['Produto'] = df_produtos['Produto_Final']
        
        if 'COD' in df_produtos.columns:
            df_produtos['COD'] = df_produtos['COD'].astype(str).str.strip().apply(
                lambda x: x[:-2] if x.endswith('.0') and x[:-2].isdigit() else x
            )
        else:
            st.error("A base de produtos deve conter uma coluna chamada 'COD'.")
            return pd.DataFrame()
            
        df_produtos['Nome_Display'] = df_produtos['Produto'].astype(str).str.strip().replace('nan', '', regex=False).str.strip()
        df_produtos['Nome_Display'] = df_produtos['Nome_Display'].mask(df_produtos['Nome_Display'] == '', 'Produto sem nome na planilha')
        
        df_produtos_validos = df_produtos[
            (df_produtos['COD'].str.len() > 0) & 
            (~df_produtos['COD'].str.lower().isin(['nan', 'none']))
        ].reset_index(drop=True)

        return df_produtos_validos
    
    except FileNotFoundError:
        st.error(f"Arquivo de produtos não encontrado em: {PRODUTOS_EXCEL}")
        return pd.DataFrame()
    except Exception as e:
        st.error(f"Erro ao carregar a base de produtos: {e}")
        return pd.DataFrame()

# Colunas da nova aba Clientes
COLUNAS_CLIENTES = ["ID_Cliente", "Nome", "Telefone", "Email", "Endereço", "Observações"]

def carregar_clientes():
    """Carrega os dados da planilha Clientes, cria a aba se não existir."""
    os.makedirs(os.path.dirname(ARQUIVO_EXCEL), exist_ok=True)
    
    dtype_force = {
        "ID_Cliente": str,
        "Nome": str,
        "Telefone": str,
        "Email": str,
        "Endereço": str,
        "Observações": str
    }

    try:
        # Tenta carregar o arquivo Excel
        xls = pd.ExcelFile(ARQUIVO_EXCEL)
        
        if "Clientes" not in xls.sheet_names:
            # Aba não existe, cria um DF vazio
            st.info("Aba 'Clientes' não encontrada. Criando uma nova.")
            df_vazio = pd.DataFrame(columns=COLUNAS_CLIENTES)
            
            # Salva a aba nova (isso requer cuidado para não apagar outras abas)
            # Vamos usar o modo 'a' (append) e 'replace' para a aba específica
            try:
                with pd.ExcelWriter(ARQUIVO_EXCEL, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                    df_vazio.to_excel(writer, sheet_name="Clientes", index=False)
            except Exception as e_write:
                # Se falhar (ex: arquivo aberto), tenta na próxima vez.
                st.warning(f"Não foi possível criar a aba 'Clientes' agora: {e_write}")
            return df_vazio
            
        # Se a aba existe, carrega ela
        df = pd.read_excel(xls, sheet_name="Clientes", dtype=dtype_force)
        
        # Garante que todas as colunas existam
        df_final = df[[col for col in COLUNAS_CLIENTES if col in df.columns]]
        for col in COLUNAS_CLIENTES:
            if col not in df_final.columns:
                df_final[col] = None
        
        return df_final[COLUNAS_CLIENTES].fillna('')

    except FileNotFoundError:
        st.error(f"Arquivo {ARQUIVO_EXCEL} não encontrado. Ele será criado na próxima vez que 'carregar_dados()' for chamado.")
        return pd.DataFrame(columns=COLUNAS_CLIENTES)
        
    except Exception as e:
        # Se o arquivo estiver corrompido ou outro erro
        st.error(f"Erro ao carregar clientes: {e}. Criando um DataFrame vazio.")
        return pd.DataFrame(columns=COLUNAS_CLIENTES)

def salvar_clientes(df):
    """Salva o DataFrame de clientes de volta na planilha Excel. Retorna True ou False."""
    try:
        df_salvar = df.copy()
        for col in COLUNAS_CLIENTES:
             if col not in df_salvar.columns:
                 df_salvar[col] = None

        # Usa 'replace' para garantir que a aba 'Clientes' seja substituída
        with pd.ExcelWriter(ARQUIVO_EXCEL, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df_salvar[COLUNAS_CLIENTES].to_excel(writer, sheet_name='Clientes', index=False)
        
        return True 
    except Exception as e:
        st.error(f"Erro ao salvar dados de clientes: {e}")
        st.error("Verifique se o arquivo BD_Loja.xlsx não está aberto em outro programa.")
        return False

# ===== FUNÇÕES DE VENDAS DO main.py (Preservadas e Modificadas) =====
def calcular_estoque():
    """Calcula o estoque atual de cada produto somando a coluna Quantidade na aba Movimento."""
    df_movimento = carregar_dados() 
    
    # Filtra as movimentações que afetam o estoque (SAÍDA é negativa, ENTRADA é positiva)
    df_estoque = df_movimento[
        df_movimento['Tipo de Movimentação'].isin(['ENTRADA', 'SAÍDA'])
    ].copy()
    
    if df_estoque.empty:
        return pd.DataFrame(columns=['COD do Produto', 'Estoque Atual'])
    
    # CORREÇÃO CRÍTICA: Garante que a Quantidade é numérica antes de somar.
    df_estoque.loc[:, 'Quantidade'] = pd.to_numeric(
        df_estoque['Quantidade'], 
        errors='coerce' # Converte não-numéricos para NaN
    ).fillna(0) # Substitui NaN por 0
    
    # Agrupa por COD do Produto e soma a Quantidade
    df_estoque_atual = df_estoque.groupby('COD do Produto')['Quantidade'].sum().reset_index()
    df_estoque_atual.rename(columns={'Quantidade': 'Estoque Atual'}, inplace=True)
    
    return df_estoque_atual[['COD do Produto', 'Estoque Atual']]


def page_products_list(df_produtos_base, df_estoque):
    st.subheader("Lista de Produtos e Controle de Estoque")
    
    # 1. Combina a lista de produtos com o estoque calculado
    df_produtos_completo = df_produtos_base.merge(
        df_estoque,
        left_on='COD', 
        right_on='COD do Produto', 
        how='left'
    )
    
    # Produtos sem movimentação terão estoque NaN, definimos como 0
    df_produtos_completo['Estoque Atual'] = df_produtos_completo['Estoque Atual'].fillna(0).astype(int)
    
    # 2. Filtro de Estoque (Padrão: Com Estoque)
    filtro_estoque = st.radio(
        "Filtrar por Status de Estoque",
        ["Com Estoque (Padrão)", "Sem Estoque", "Todos"],
        index=0,
        horizontal=True
    )
    
    df_filtrado = df_produtos_completo.copy()

    if filtro_estoque == "Com Estoque (Padrão)":
        df_filtrado = df_filtrado[df_filtrado['Estoque Atual'] > 0]
    elif filtro_estoque == "Sem Estoque":
        df_filtrado = df_filtrado[df_filtrado['Estoque Atual'] <= 0]
    
    st.markdown(f"**Total de Produtos Exibidos:** {len(df_filtrado)}")
    
    # 3. Exibição da Tabela
    # 'Produto' é a coluna com o PROCV/CONCATENAR no seu Excel
    cols_display = [
        'COD', 'Produto', 'Estoque Atual', 'Preço Venda', 'Preço Custo', 
        'Marca', 'Categoria', 'Coleção' 
    ]
    
    cols_final = [col for col in cols_display if col in df_filtrado.columns]
    df_display = df_filtrado[cols_final].copy()
    
    df_display.rename(columns={'COD': 'COD do Produto', 'Produto': 'Nome Produto'}, inplace=True)
    
    # Formatação de preços
    if not df_display.empty: 
        for col in ['Preço Venda', 'Preço Custo']:
            if col in df_display.columns:
                df_display[col] = pd.to_numeric(df_display[col], errors='coerce').fillna(0).map('R$ {:,.2f}'.format).str.replace(',', 'X').str.replace('.', ',').str.replace('X', '.')
    
    st.dataframe(df_display, use_container_width=True, hide_index=True)


def registrar_venda():
    """Interface para registro de uma nova venda (saída de estoque)."""
    st.title("Registrar Nova Venda")
    
    if "carrinho" not in st.session_state:
        st.session_state["carrinho"] = []

    df_produtos_validos = carregar_produtos()
    df_clientes = carregar_clientes() 

    if df_produtos_validos.empty:
        st.warning("Não foi possível carregar a base de produtos válidos.")
        return
    
    opcoes_produtos = (df_produtos_validos['COD'] + " - " + df_produtos_validos['Marca'] + " "+df_produtos_validos['Produto']).tolist()
    opcoes_produtos.insert(0, "Selecione um produto...")
    
    st.subheader("1. Adicionar Item")
    
    col1, col2, col3, col4 = st.columns([3, 1.5, 1.5, 1.5])
    
    with col1:
        produto_selecionado_add = st.selectbox(
            "Produto", 
            opcoes_produtos,
            key='selectbox_produto_add' 
        )

    produto_info = None
    cod_selecionado = None
    preco_custo_unitario_default = 0.0
    preco_venda_unitario_default = 0.0
    
    if produto_selecionado_add != "Selecione um produto...":
        cod_selecionado = produto_selecionado_add.split(" - ")[0].strip()
        produto_info_df = df_produtos_validos[df_produtos_validos['COD'] == cod_selecionado]
        if not produto_info_df.empty:
            produto_info = produto_info_df.iloc[0]
            try:
                preco_custo_default = pd.to_numeric(produto_info.get('Preço Custo', 0), errors='coerce')
                preco_venda_default = pd.to_numeric(produto_info.get('Preço Venda', 0), errors='coerce')
                preco_custo_unitario_default = float(preco_custo_default) if pd.notna(preco_custo_default) else 0.0
                preco_venda_unitario_default = float(preco_venda_default) if pd.notna(preco_venda_default) else 0.0
            except Exception:
                preco_custo_unitario_default = 0.0
                preco_venda_unitario_default = 0.0

    with col2:
        qtd_add = st.number_input("Quantidade", min_value=1, value=1, step=1, key='input_quantidade_add')
    with col3:
        preco_custo_unitario_add = st.number_input(
            "Preço Custo Unit.", 
            min_value=0.0, 
            value=preco_custo_unitario_default, 
            format="%.2f",
            key='input_custo_add'
        )
    with col4:
        preco_venda_unitario_add = st.number_input(
            "Preço Venda Unit.", 
            min_value=0.0, 
            value=preco_venda_unitario_default, 
            format="%.2f",
            key='input_venda_add'
        )

    if st.button("Adicionar Item ao Carrinho", key='btn_add_carrinho'):
        if produto_selecionado_add == "Selecione um produto...":
            st.error("Selecione um produto.")
        elif cod_selecionado is None:
            st.error("Erro ao carregar informações do produto.")
        elif qtd_add <= 0:
            st.error("A quantidade deve ser maior que zero.")
        elif preco_venda_unitario_add <= 0:
            st.warning("O preço de venda unitário é zero. Por favor, corrija.")
        else:
            if produto_info is not None:
                marca = produto_info.get('Marca', '').strip()
                nome_base = produto_info.get('Produto', '').strip()
                
                if marca:
                    nome_prod = f"{marca} {nome_base}"
                else:
                    nome_prod = nome_base
            else:
                nome_prod = produto_selecionado_add.split(" - ", 1)[1] 
            
            item = {
                "COD do Produto": cod_selecionado,
                "Produto": nome_prod, 
                "Quantidade": qtd_add,
                "Preço Custo Unitário": preco_custo_unitario_add,
                "Preço Venda Unitário": preco_venda_unitario_add,
                "Preço Custo Total": qtd_add * preco_custo_unitario_add,
                "Preço Venda Total": qtd_add * preco_venda_unitario_add,
            }
            st.session_state["carrinho"].append(item)
            st.success(f"{qtd_add}x {item['Produto']} adicionado(s) ao carrinho.")
            time.sleep(2)
            st.rerun() 
            return

    st.markdown("---")
    
    st.subheader("2. Resumo da Venda e Pagamento")

    if not st.session_state.get("carrinho"):
        st.info("O carrinho está vazio. Adicione um produto para continuar.")
        return

    df_carrinho = pd.DataFrame(st.session_state["carrinho"])
    
    df_display = df_carrinho[[
        "Produto", "Quantidade", "Preço Venda Unitário", "Preço Venda Total"
    ]].copy()
    for col in ["Preço Venda Unitário", "Preço Venda Total"]:
        df_display[col] = df_display[col].apply(lambda x: f"R$ {x:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))

    st.dataframe(df_display, use_container_width=True)
    
    if st.button("Limpar Carrinho", key='btn_limpar_carrinho'):
        st.session_state["carrinho"] = []
        st.rerun()
        return

    total_venda = df_carrinho["Preço Venda Total"].sum()
    st.markdown(f"#### **Total da Venda: R$ {total_venda:,.2f}**".replace(",", "X").replace(".", ",").replace("X", "."))

    data_prevista_vencimento = None
    parcelas_info = []
    
    tipo_pagamento_status = st.selectbox(
        "Status da Venda",
        ["À Vista (Pago)", "Fiado (A Receber)", "Parcelado (A Receber)"],
        key="tipo_pagamento_key_final"
    )

    if tipo_pagamento_status == "À Vista (Pago)":
        meio_pagamento = st.radio(
            "Meio de Pagamento",
            ["Pix", "Cartão", "Dinheiro"],
            key="meio_pagamento_radio",
            horizontal=True
        )
    elif tipo_pagamento_status == "Fiado (A Receber)":
        data_prevista_vencimento = st.date_input(
            "Data Prevista de Recebimento",
            datetime.today(),
            key="data_prevista_fiado_final"
        ).strftime('%Y-%m-%d')
        meio_pagamento = "Fiado" 
    elif tipo_pagamento_status == "Parcelado (A Receber)":
        col_parcelas_1, col_parcelas_2 = st.columns(2)
        with col_parcelas_1:
            num_parcelas = st.number_input(
                "Número de Parcelas", 
                min_value=2, max_value=12, 
                value=st.session_state.get('num_parcelas_key_final', 2), 
                step=1, 
                key="num_parcelas_key_final" 
            )
        valor_parcela = total_venda / num_parcelas
        st.markdown(f"**Valor por Parcela:** R$ {valor_parcela:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))
        with col_parcelas_2:
            primeira_data = st.date_input(
                "Data da 1ª Parcela", 
                value=st.session_state.get('data_primeira_parcela_final_val', datetime.today()), 
                key="data_primeira_parcela_final_val"
            )
        for i in range(num_parcelas):
            data_parcela = (primeira_data + DateOffset(months=i)).strftime('%Y-%m-%d')
            parcelas_info.append({
                "Parcela": f"{i+1}/{num_parcelas}",
                "Data Prevista": data_parcela,
                "Valor Parcela": round(valor_parcela, 2)
            })
        st.write("📅 **Resumo das Parcelas:**")
        st.dataframe(pd.DataFrame(parcelas_info), use_container_width=True)
        meio_pagamento = "Parcelado" 

    # ⚠️ MUDANÇA PRINCIPAL AQUI ⚠️
    st.subheader("3. Cliente e Observações")

    # O Selectbox do Cliente agora fica FORA do formulário
    # Isso força o Streamlit a recarregar a página quando ele muda
    opcoes_clientes_venda = ["Selecione um cliente cadastrado..."] + \
                            sorted(df_clientes['Nome'].unique().tolist()) + \
                            ["+ Cadastrar novo cliente na hora"]
    
    cliente_selecao = st.selectbox(
        "Cliente *", 
        opcoes_clientes_venda, 
        key='cliente_venda_select_form' # A chave pode ser a mesma
    )
    # ⚠️ FIM DA MUDANÇA ⚠️
    
    with st.form("form_finalizar_venda"):
        
        # O campo de texto para o novo cliente fica DENTRO do formulário
        # Mas ele só aparece se o selectbox (agora fora) estiver com o valor certo
        if cliente_selecao == "+ Cadastrar novo cliente na hora":
            st.text_input(
                "Nome do Novo Cliente *", 
                key='input_cliente_final_form_novo' # Chave de estado para o novo nome
            )
        
        Observacoes_final = st.text_area("Observações (opcional)", value="", key="Observacoes_key_final_form")
        submitted = st.form_submit_button("Finalizar Venda e Registrar Movimento")
        
        if submitted:
            # A lógica de leitura dos valores no submit permanece a mesma
            
            # Pega o valor do selectbox (de fora)
            cliente_selecao_valor = st.session_state.get('cliente_venda_select_form')
            
            cliente_final = ""
            if cliente_selecao_valor == "+ Cadastrar novo cliente na hora":
                # Pega o valor do text_input (de dentro)
                cliente_final = st.session_state.get('input_cliente_final_form_novo', "").strip()
            elif cliente_selecao_valor != "Selecione um cliente cadastrado...":
                cliente_final = cliente_selecao_valor
            
            # Validação
            if not cliente_final:
                st.error("O campo Cliente é obrigatório. Selecione um ou cadastre um novo.")
            elif total_venda <= 0: 
                st.error("O Valor Total de Venda deve ser maior que zero.")
            else:
                
                # Lógica para salvar novo cliente
                if cliente_selecao_valor == "+ Cadastrar novo cliente na hora":
                    df_clientes_check = carregar_clientes()
                    
                    if cliente_final.lower() not in df_clientes_check['Nome'].str.strip().str.lower().values:
                        st.info(f"Registrando novo cliente: '{cliente_final}'...")
                        novo_id = f"C-{uuid.uuid4().hex[:6].upper()}"
                        novo_cliente_data = {
                            "ID_Cliente": novo_id, "Nome": cliente_final, "Telefone": "", 
                            "Email": "", "Endereço": "", "Observações": "Cadastrado via Venda"
                        }
                        df_clientes_novo = pd.concat([df_clientes_check, pd.DataFrame([novo_cliente_data])], ignore_index=True)
                        
                        if not salvar_clientes(df_clientes_novo):
                            st.error("Erro ao salvar o novo cliente na aba 'Clientes'. A venda foi cancelada. Verifique o arquivo Excel.")
                            return 
                        
                        st.success(f"Novo cliente '{cliente_final}' cadastrado com sucesso.")
                    else:
                        st.warning(f"O cliente '{cliente_final}' já existia no cadastro. Registrando a venda para ele.")

                # --- Continua com o registro da venda ---
                
                df = carregar_dados() 
                
                id_venda = f"V-{uuid.uuid4().hex[:6].upper()}"
                status_venda = "PAGO" if tipo_pagamento_status == "À Vista (Pago)" else "A RECEBER"
                data_registro = datetime.now() 
                registros = []
                
                if tipo_pagamento_status == "Parcelado (A Receber)":
                    obs_adicional = Observacoes_final.strip()
                    for item in st.session_state["carrinho"]:
                        for i, p in enumerate(parcelas_info):
                            is_last_installment = (i == num_parcelas - 1)
                            
                            quantidade_registro = -item["Quantidade"] if is_last_installment else 0
                            custo_registro = -item["Preço Custo Total"] if is_last_installment else 0
                            
                            obs_parcela = f"Parcela {p['Parcela']} | {obs_adicional}" if obs_adicional else f"Parcela {p['Parcela']}"
                            data_vencimento_full = f"{p['Data Prevista']} 00:00:00"
                            
                            registros.append({
                                "Data": data_registro,
                                "COD do Produto": item["COD do Produto"], 
                                "Produto": item["Produto"], 
                                "Cliente": cliente_final, 
                                "Tipo de Movimentação": "SAÍDA",
                                "Quantidade": quantidade_registro, 
                                "Preço Custo Total": custo_registro, 
                                "Preço Venda Total": p["Valor Parcela"], 
                                "Observações": obs_parcela,
                                "Status": "A RECEBER",
                                "Data Prevista": data_vencimento_full,
                                "Tipo de Pagamento": meio_pagamento,
                                "ID_Venda": id_venda,
                            })
                else:
                    if tipo_pagamento_status == "À Vista (Pago)":
                        data_final_pgto_venc = data_registro.strftime('%Y-%m-%d %H:%M:%S')
                    else: 
                        if data_prevista_vencimento:
                            data_final_pgto_venc = f"{data_prevista_vencimento} 00:00:00"
                        else:
                            data_final_pgto_venc = np.nan

                    for item in st.session_state["carrinho"]:
                        registros.append({
                            "Data": data_registro,
                            "COD do Produto": item["COD do Produto"],
                            "Produto": item["Produto"],
                            "Cliente": cliente_final, 
                            "Tipo de Movimentação": "SAÍDA",
                            "Quantidade": -item["Quantidade"], 
                            "Preço Custo Total": -item["Preço Custo Total"], 
                            "Preço Venda Total": item["Preço Venda Total"], 
                            "Observações": Observacoes_final.strip(), 
                            "Status": status_venda,
                            "Data Prevista": data_final_pgto_venc,
                            "Tipo de Pagamento": meio_pagamento, 
                            "ID_Venda": id_venda,
                        })

                df_novo = pd.concat([df, pd.DataFrame(registros)], ignore_index=True)
                
                save_successful = salvar_dados(df_novo)  
                
                if save_successful: 
                    st.session_state["carrinho"] = [] 
                    st.success(f"Venda {id_venda} registrada com sucesso para o cliente '{cliente_final}'!")
                    time.sleep(2) 
                    st.rerun() 
                else:
                    st.error("Falha ao registrar venda. Verifique as mensagens de erro acima.")

def mostrar_saldos(): 
# ... (Restante do código mantido)
    """Calcula e exibe o saldo total a receber por cliente."""
    st.title("Saldo de Clientes (Contas a Receber)")
    df = carregar_dados()

    df_a_receber = df[(df['Tipo de Movimentação'] == 'SAÍDA') & (df['Status'] == 'A RECEBER')].copy()

    if df_a_receber.empty:
        st.info("Não há saldos pendentes a receber de clientes.")
        return

    df_saldo_efetivo = df_a_receber[df_a_receber['Preço Venda Total'] > 0]
    saldo_clientes = df_saldo_efetivo.groupby('Cliente')['Preço Venda Total'].sum().reset_index()
    saldo_clientes.columns = ['Cliente', 'Total a Receber']

    saldo_clientes['Total a Receber'] = saldo_clientes['Total a Receber'].map('{:,.2f}'.format).str.replace(',', 'X').str.replace('.', ',').str.replace('X', '.')
    
    st.subheader("Resumo do Saldo Total por Cliente")
    st.dataframe(saldo_clientes, use_container_width=True, hide_index=True)
    
    st.subheader("Detalhes dos Itens em Aberto")
    df_detalhe = df_a_receber[['Data', 'ID_Venda', 'Cliente', 'Produto', 'Preço Venda Total', 'Data Prevista', 'Observações']].copy()
    df_detalhe.rename(columns={'Preço Venda Total': 'Valor a Receber', 'Data Prevista': 'Data Pgto'}, inplace=True)
    
    df_detalhe['Valor a Receber'] = df_detalhe['Valor a Receber'].map('{:,.2f}'.format).str.replace(',', 'X').str.replace('.', ',').str.replace('X', '.')
    
    df_detalhe['Data Pgto'] = df_detalhe['Data Pgto'].apply(lambda x: x.strftime('%d/%m/%Y') if pd.notna(x) else 'N/A')
    
    st.dataframe(df_detalhe.sort_values(by='Data Pgto', ascending=True).reset_index(drop=True), use_container_width=True, hide_index=True)

def atualizar_celula_excel(cod_produto, coluna_nome, novo_valor):
    """
    Atualiza uma célula específica (Ex: Preço Custo) na aba 'Produtos' usando openpyxl.
    Preserva todas as outras formatações e design.
    """
    try:
        # 1. Carregar o Workbook
        book = openpyxl.load_workbook(PRODUTOS_EXCEL)
        sheet = book["Produtos"]
        
        COLUMNS_MAP = {
            'Preço Custo': 'K', 
            'Preço Venda': 'L',
            'COD': 'A'
        }
        
        coluna_letra = COLUMNS_MAP.get(coluna_nome)
        if not coluna_letra:
            st.error(f"Coluna {coluna_nome} não mapeada para atualização.")
            return False
            
        # Garante que o COD de busca esteja limpo (string e caixa alta)
        cod_busca = str(cod_produto).strip().upper()

        # 2. Encontrar a linha
        # Assumindo que o COD está na Coluna A (primeiro índice '0')
        row_to_update = -1
        
        # Iterar a partir da linha 2 (índice 0 corresponde à linha 2)
        for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True)):
            excel_cod_raw = row[0]
            
            # ⚠️ CORREÇÃO CHAVE: Converter o valor do Excel para string, remover espaços e forçar caixa alta
            # Isso resolve o problema de comparação entre int(1) e str('1') ou espaços ocultos.
            if excel_cod_raw is not None:
                excel_cod_limpo = str(excel_cod_raw).strip().upper()
                
                if excel_cod_limpo == cod_busca: 
                    # i é o índice na iteração (0 para L2, 1 para L3, etc.)
                    row_to_update = i + 2 
                    break
        
        if row_to_update == -1:
            st.warning(f"Produto com COD '{cod_produto}' não encontrado no Excel. Verifique a coluna 'COD' na sua planilha.")
            return False

        # 3. Atualizar a célula específica
        cell_ref = coluna_letra + str(row_to_update)
        sheet[cell_ref].value = novo_valor # Escreve o novo valor

        # 4. Salvar o Workbook
        book.save(PRODUTOS_EXCEL)
        
        return True
        
    except Exception as e:
        st.error(f"Erro ao atualizar célula no Excel: {e}")
        st.error("Verifique se o arquivo BD_Loja.xlsx não está aberto em outro programa.")
        return False
    
def atualizar_produto(df_produtos_base):
    '''Interface para atualizar o preço de custo ou venda de um produto existente.'''
    st.subheader("Atualizar Preçode Produto Existente")

    # 1. Preparar dados
    df_produtos_base['Preço Venda_float'] = pd.to_numeric(
        df_produtos_base['Preço Venda'], errors='coerce'
    ).fillna(0.0) 
    
    opcoes_produto = df_produtos_base.apply(
        # Usa a nova coluna numérica para a formatação com :.2f
        lambda row: f"{row['COD']} - {row['Produto']} (R$ {row['Preço Venda_float']:.2f})", 
        axis=1
    ).tolist()
    
    # 2. Formulário de Seleção e Atualização
    with st.form("form_atualizar_produto"):
        produto_selecionado_str = st.selectbox(
            "Selecione o Produto para Atualizar",
            opcoes_produto,
            help="O nome exibido é o resultado da fórmula do seu Excel."
        )
        
        cod_selecionado = produto_selecionado_str.split(' - ')[0]
        produto_atual = df_produtos_base[df_produtos_base['COD'] == cod_selecionado].iloc[0]
        
        st.markdown(f"#### Produto Selecionado: **{produto_atual['Produto']}**")
        
        col1, col2 = st.columns(2)
        
        # Garante que os valores lidos do Excel sejam floats
        try:
            preco_custo_atual = float(produto_atual['Preço Custo'])
            preco_venda_atual = float(produto_atual['Preço Venda'])
        except (ValueError, TypeError):
            st.error("Erro de dado: O 'Preço Custo' ou 'Preço Venda' está com valor inválido no Excel.")
            return

        with col1:
            novo_custo = st.number_input(
                "Novo Preço Custo (Atual: R$ {:.2f})".format(preco_custo_atual), 
                min_value=0.0, 
                value=preco_custo_atual, 
                format="%.2f"
            )
        with col2:
            novo_venda = st.number_input(
                "Novo Preço Venda (Atual: R$ {:.2f})".format(preco_venda_atual), 
                min_value=0.0, 
                value=preco_venda_atual, 
                format="%.2f"
            )

        submitted = st.form_submit_button("Atualizar Preços")

        if submitted:
            # 3. Executar a Atualização Direta no Excel
            
            # Checa se houve mudanças para evitar salvar sem necessidade
            custo_mudou = novo_custo != preco_custo_atual
            venda_mudou = novo_venda != preco_venda_atual
            
            success_custo = True
            success_venda = True

            if custo_mudou:
                success_custo = atualizar_celula_excel(cod_selecionado, 'Preço Custo', novo_custo)
            
            if venda_mudou:
                success_venda = atualizar_celula_excel(cod_selecionado, 'Preço Venda', novo_venda)
                
            if success_custo and success_venda:
                st.success(f"Preços do produto {cod_selecionado} atualizados com sucesso! (Apenas as colunas K e L foram modificadas)")
                time.sleep(1)
                st.rerun()
            else:
                st.error("Falha ao atualizar no Excel. Verifique a mensagem de erro acima.")

def atualizar_recebimento():
    """Interface para atualizar um item 'A RECEBER' para 'PAGO'."""
    st.title("Atualizar Recebimento")
    df = carregar_dados()

    df_pendente = df[(df['Tipo de Movimentação'] == 'SAÍDA') & (df['Status'] == 'A RECEBER')].copy()
    
    if df_pendente.empty:
        st.info("Não há recebimentos pendentes para atualizar.")
        return

    df_pendente.rename(columns={'Preço Venda Total': 'Valor a Receber', 'Data Prevista': 'Data Pgto'}, inplace=True)
    
    df_pendente['Data Pgto Display'] = df_pendente['Data Pgto'].apply(lambda x: x.strftime('%d/%m/%Y') if pd.notna(x) else 'N/A')
    
    df_display_pendente = df_pendente[['ID_Venda', 'Cliente', 'Produto', 'Valor a Receber', 'Data Pgto Display', 'Observações']].reset_index()
    df_display_pendente.columns = ['Índice', 'ID_Venda', 'Cliente', 'Produto', 'Valor a Receber', 'Data Pgto', 'Observações']
    
    df_display_pendente['Valor a Receber'] = df_display_pendente['Valor a Receber'].map('{:,.2f}'.format).str.replace(',', 'X').str.replace('.', ',').str.replace('X', '.')
    
    st.dataframe(df_display_pendente.drop(columns='Índice'), use_container_width=True, hide_index=True)

    opcoes_selecao = [
        f"Índice {row['Índice']} | ID: {row['ID_Venda']} | Cliente: {row['Cliente']} | Produto: {row['Produto']} | Valor: {row['Valor a Receber']} | Previsto: {row['Data Pgto']}" 
        for i, row in df_display_pendente.iterrows()
    ]
    opcoes_selecao.insert(0, "Selecione o Recebimento para Marcar como Pago...")

    registro_selecionado = st.selectbox("Selecione o Recebimento a ser Atualizado para 'PAGO'", opcoes_selecao, key='select_recebimento_pagar')

    if registro_selecionado != "Selecione o Recebimento para Marcar como Pago...":
        index_selecionado = int(registro_selecionado.split(" | ")[0].replace("Índice ", ""))
        
        if st.button(f"Confirmar Pagamento do Registro (Índice {index_selecionado})", key='btn_confirmar_pagamento'):
            dfn = df.copy()
            
            # Altera o Status
            if 'Status' not in dfn.columns: dfn['Status'] = None
            dfn.loc[index_selecionado, 'Status'] = 'PAGO'
            
            # Registra a data atual (com hora) na coluna Data Prevista
            if 'Data Prevista' not in dfn.columns: dfn['Data Prevista'] = None
            dfn.loc[index_selecionado, 'Data Prevista'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            
            # <<< 7. CAPTURA O STATUS DO SALVAMENTO NO RECEBIMENTO
            save_successful = salvar_dados(dfn)
            
            if save_successful: # <<< 8. VERIFICA SE FOI SUCESSO ANTES DE PROSSEGUIR
                st.success(f"Recebimento do registro {index_selecionado} atualizado para PAGO.")
                time.sleep(2) # <<< 9. ADICIONA O TIMER
                st.rerun()
            else:
                st.error("Falha ao atualizar recebimento. Verifique as mensagens de erro acima.")

def mostrar_todas_vendas():
    """Exibe um histórico de todas as movimentações."""
# ... (Restante do código mantido)
    st.title("Histórico Completo de Movimentações")
    df = carregar_dados()
    
    if df.empty:
        st.info("Nenhuma movimentação registrada.")
        return

    df_display = df.copy()
    
    for col in ['Preço Custo Total', 'Preço Venda Total']:
        if col in df_display.columns:
            df_display[col] = df_display[col].fillna(0).map('{:,.2f}'.format).str.replace(',', 'X').str.replace('.', ',').str.replace('X', '.')
    
    df_display.rename(columns={'Data Prevista': 'Data Pgto'}, inplace=True)

    def format_pgto_date(row):
        dt = row['Data Pgto']
        if pd.isna(dt) or dt is None:
            return ''
        
        # A data é tratada como datetime pelo Pandas ao carregar.
        # Precisamos diferenciar se é Data de Vencimento (A RECEBER) ou Data de Pagamento (PAGO)
        if row['Status'] == 'PAGO':
            # Data de Pagamento: queremos data e hora
            return dt.strftime('%d/%m/%Y %H:%M')
        elif row['Status'] == 'A RECEBER':
            # Data de Vencimento: queremos apenas a data
            return dt.strftime('%d/%m/%Y')
        return ''

    if "Data Pgto" in df_display.columns:
        df_display["Data Pgto"] = df_display.apply(format_pgto_date, axis=1)

    col_order = [
        "Data", "ID_Venda", "Cliente", "Produto", "Tipo de Movimentação",
        "Quantidade", "Preço Custo Total", "Preço Venda Total", "Status",
        "Data Pgto", "Observações"
    ]
    
    cols_final = [col for col in col_order if col in df_display.columns]
    df_display = df_display[cols_final]
    
    if "Data" in df_display.columns:
        df_display["Data"] = pd.to_datetime(df_display["Data"], errors="coerce")
        df_display = df_display.sort_values(by="Data", ascending=False).reset_index(drop=True)
    
    st.dataframe(df_display, use_container_width=True, hide_index=True)


# ===== FUNÇÕES DE DASHBOARD/DESIGN (CORRIGIDO E MELHORADO) =====

def inject_css():
# ... (Restante do código mantido)
    """
    Injeta o CSS para o tema 'Superstore' e customiza Selectbox, Botões e Inputs 
    com o esquema de cores rosa harmonioso.
    """
    # Cores do tema: #880e4f (Escuro), #e91e63 (Choque/Accent), #fce4ec (Claro), #ffebee (Mais claro ainda, para fundo de campo)
    st.markdown("""
    <style>
    /* Estilos de Tema Geral */
    .st-emotion-cache-1r6r0k9 {padding-top: 1rem;}
    .st-emotion-cache-1d387o9 { /* Barra Lateral */ 
        background-color: #fce4ec; /* Rosa Suave */ 
    } 
    .st-emotion-cache-1cpx922 { /* Botões de Rádio (Menu) */
        color: #880e4f; /* Rosa Escuro */ 
    } 
    .big-title { 
        font-size: 2.5em; 
        color: #880e4f; 
        font-weight: bold; 
        margin-bottom: 0px; 
    } 
    .subtitle { 
        font-size: 1.1em; 
        color: #e91e63; /* Rosa Choque */ 
        margin-bottom: 20px; 
    }
    
    /* === ESTILOS PARA INPUTS (st.text_input, st.number_input, st.date_input, st.text_area) === */
    /* Target o container do input (Text, Number, Date, Area) */
    [data-baseweb="input"], [data-baseweb="textarea"], [data-baseweb="select"] {
        border-radius: 8px;
        border: 1px solid #e91e63 !important; /* Borda rosa choque */
        background-color: #ffebee !important; /* Fundo ROSA CLARO SOLICITADO */
    }
    /* Target o texto digitado dentro do input */
    [data-baseweb="input"] input, [data-baseweb="textarea"] textarea {
        color: #880e4f !important; /* Texto em rosa escuro */
        background-color: #ffebee !important; /* Garante que o fundo interno também seja rosa claro */
    }

    /* === ESTILOS PARA BOTÕES (st.button, st.form_submit_button) === */
    .stButton>button, [data-testid="stFormSubmitButton"]>button {
        background-color: #e91e63 !important; /* Rosa Choque */
        color: white !important;
        border-radius: 8px;
        border: 1px solid #880e4f; /* Borda Rosa Escuro */
        font-weight: bold;
        padding: 8px 16px;
        transition: background-color 0.3s;
    }
    .stButton>button:hover, [data-testid="stFormSubmitButton"]>button:hover {
        background-color: #880e4f !important; /* Rosa Escuro no Hover */
        color: white !important;
        border: 1px solid #e91e63;
    }
    
    /* === ESTILOS PARA SELECTBOX (Dropdowns) === */
    /* Campo de exibição principal */
    [data-testid="stSelectbox"] div[data-baseweb="select"] > div:first-child {
        background-color: #ffebee !important; 
        color: #880e4f !important; 
        border: 1px solid #e91e63 !important; 
    }
    /* Seta do dropdown */
    [data-testid="stSelectbox"] svg {
        color: #880e4f !important; 
        fill: #880e4f !important;
    }
    /* Lista suspensa (o dropdown que abre) */
    div[role="listbox"] {
        background-color: #fce4ec !important; 
        border: 1px solid #e91e63;
    }
    /* Opção na lista ao passar o mouse (hover) */
    div[role="option"]:hover {
        background-color: #e91e63 !important; 
        color: white !important; 
    }
    /* Opção selecionada na lista */
    div[role="option"][aria-selected="true"] {
        background-color: #880e4f !important; 
        color: white !important;
    }
    
    </style>
    """, unsafe_allow_html=True)


# ===== LÓGICA DO DASHBOARD (CORRIGIDO STATUS E ADICIONADO FILTROS DE PRODUTO) =====

def page_dashboard_logic(df_movimento, df_produtos):
# ... (Restante do código mantido)
    df_movimento = df_movimento.copy()
    
    # 1. Filtra apenas SAÍDAS (Vendas) e Preço Venda Total > 0 (Base geral)
    # Esta base inclui PAGO e A RECEBER
    df_vendas_base_all = df_movimento[
        (df_movimento['Tipo de Movimentação'] == 'SAÍDA') & 
        (df_movimento['Preço Venda Total'] > 0)
    ].copy()
    
    if df_vendas_base_all.empty:
        st.warning("Não há dados de vendas válidas (SAÍDA com Preço Venda Total > 0) para o Dashboard.")
        return
        
    # Certifica-se que a coluna 'Data' é datetime
    df_vendas_base_all['Data'] = pd.to_datetime(df_vendas_base_all['Data'], errors='coerce')
    df_vendas_base_all = df_vendas_base_all.dropna(subset=['Data'])
    
    min_date = df_vendas_base_all['Data'].min().date() if not df_vendas_base_all.empty else date.today()
    max_date_data = df_vendas_base_all['Data'].max().date() if not df_vendas_base_all.empty else date.today() 

    # --- NOVO FILTRO DE PERÍODO (SUBSTITUI OS DOIS COLUMNS DE DATA) ---
    st.subheader("Filtros de Período e Produtos")
    
    # Ajustando o layout: 2 para o novo filtro de data/período, e 1.5, 1.5, 1.5 para os filtros de produto
    col_filtro_periodo, col_filtro_3, col_filtro_4, col_filtro_5 = st.columns([2, 1.5, 1.5, 1.5]) 
    
    # Inicializa as datas com o intervalo "Todo o Período" (será sobrescrito)
    data_inicio_selecionada = min_date
    data_fim_selecionada = max_date_data
    
    with col_filtro_periodo:
        filtro_periodo = st.selectbox(
            "Selecione o Período de Análise",
            [
                "Últimos 30 Dias",
                "Últimos 7 Dias",
                "Últimos 365 Dias",
                "Todo o Período",
                "Intervalo Personalizado"
            ],
            index=0, # Inicia em 30 dias
            key='select_filtro_periodo'
        )

    # Lógica de cálculo da data baseada na seleção
    if filtro_periodo == "Últimos 7 Dias":
        data_inicio_selecionada = max(min_date, max_date_data - pd.Timedelta(days=6))
        data_fim_selecionada = max_date_data
    elif filtro_periodo == "Últimos 30 Dias":
        # De 30 dias atrás até o último registro
        data_inicio_selecionada = max(min_date, max_date_data - pd.Timedelta(days=29))
        data_fim_selecionada = max_date_data
    elif filtro_periodo == "Últimos 365 Dias":
        # De 365 dias atrás até o último registro
        data_inicio_selecionada = max(min_date, max_date_data - pd.Timedelta(days=364))
        data_fim_selecionada = max_date_data
    elif filtro_periodo == "Todo o Período":
        data_inicio_selecionada = min_date
        data_fim_selecionada = max_date_data
    
    # Se for "Intervalo Personalizado", permite ao usuário selecionar as datas
    if filtro_periodo == "Intervalo Personalizado":
        # Usamos o espaço da coluna de filtro de dados para os inputs de data
        col_pers_1, col_pers_2 = col_filtro_periodo.columns(2)
        with col_pers_1:
            data_inicio_input = st.date_input(
                "Data Inicial Personalizada", 
                min_value=min_date, 
                max_value=max_date_data, 
                value=data_inicio_selecionada, 
                key='data_inicio_personalizada'
            )
        with col_pers_2:
            data_fim_input = st.date_input(
                "Data Final Personalizada", 
                min_value=min_date, 
                max_value=max_date_data, 
                value=data_fim_selecionada, 
                key='data_fim_personalizada'
            )
        # Atualiza as datas a serem usadas no filtro
        data_inicio_selecionada = data_inicio_input
        data_fim_selecionada = data_fim_input

    # Converte as datas selecionadas/calculadas para datetime (inclui o dia todo na data final)
    dt_inicio = pd.to_datetime(data_inicio_selecionada)
    dt_fim = pd.to_datetime(data_fim_selecionada) + pd.Timedelta(days=1, seconds=-1) 

    # Aplica filtro de Data na base ALL-STATUS
    df_filtrado_data_all = df_vendas_base_all[
        (df_vendas_base_all['Data'] >= dt_inicio) & 
        (df_vendas_base_all['Data'] <= dt_fim)
    ].copy()

    if df_filtrado_data_all.empty:
        st.warning(f"Nenhuma venda encontrada no período de {data_inicio_selecionada.strftime('%d/%m/%Y')} a {data_fim_selecionada.strftime('%d/%m/%Y')}.")
        return

    df_filtrado_data_all['Lucro Bruto'] = df_filtrado_data_all['Preço Venda Total'] + df_filtrado_data_all['Preço Custo Total']

    # 2. Mescla com a base de produtos para obter Categoria, Coleção e Marca
    cols_to_merge = ['COD', 'Categoria', 'Coleção', 'Marca']
    available_cols = [col for col in cols_to_merge if col in df_produtos.columns]
    
    # DataFrame Base para análise (ALL-STATUS, filtrado por Data)
    df_analise_base_all = df_filtrado_data_all.merge(
        df_produtos[available_cols].drop_duplicates(subset=['COD']), 
        left_on='COD do Produto', 
        right_on='COD', 
        how='left'
    )
    
    # Preenche colunas de produto desconhecido
    fillna_dict = {}
    if 'Categoria' in df_analise_base_all.columns: fillna_dict['Categoria'] = 'Desconhecida'
    if 'Coleção' in df_analise_base_all.columns: fillna_dict['Coleção'] = 'Desconhecida'
    # 'Marca' nem sempre existe, então só preenche se for carregada
    if 'Marca' in df_analise_base_all.columns: fillna_dict['Marca'] = 'Desconhecida' 

    df_analise_base_all.fillna(fillna_dict, inplace=True)
    
    # 3. Criação e Aplicação dos Novos Filtros de Produto
    
    # Obtém opções únicas do DataFrame (dentro do período)
    categorias = ['Todas'] + sorted(df_analise_base_all['Categoria'].unique().tolist())
    colecoes = ['Todas'] + sorted(df_analise_base_all['Coleção'].unique().tolist())
    marcas = ['Todas'] + sorted(df_analise_base_all['Marca'].unique().tolist()) if 'Marca' in df_analise_base_all.columns else ['Todas']

    with col_filtro_3:
        filtro_categoria = st.selectbox("Filtrar por Categoria", categorias)
    with col_filtro_4:
        filtro_colecao = st.selectbox("Filtrar por Coleção", colecoes)
    with col_filtro_5:
        # Se 'Marca' não existir, o selectbox mostrará apenas 'Todas'
        filtro_marca = st.selectbox("Filtrar por Marca", marcas)

    # Aplica os filtros de produto ao DataFrame ALL-STATUS para contagem de Total de Vendas
    df_analise_all = df_analise_base_all.copy()
    
    if filtro_categoria != 'Todas':
        df_analise_all = df_analise_all[df_analise_all['Categoria'] == filtro_categoria]
        
    if filtro_colecao != 'Todas':
        df_analise_all = df_analise_all[df_analise_all['Coleção'] == filtro_colecao]
        
    if filtro_marca != 'Todas' and 'Marca' in df_analise_all.columns:
        df_analise_all = df_analise_all[df_analise_all['Marca'] == filtro_marca]
        
    if df_analise_all.empty:
        st.warning("Nenhuma venda encontrada com a combinação de filtros selecionada.")
        return

    # *** CÁLCULO DO KPI TOTAL DE VENDAS (NÚMERO DE TRANSAÇÕES, INDEPENDENTE DO STATUS) ***
    # Usa a base ALL-STATUS (df_analise_all)
    total_vendas = df_analise_all['ID_Venda'].nunique()


    # 4. Filtra por STATUS = 'PAGO' para todas as métricas financeiras e gráficos
    df_analise_pago = df_analise_all[df_analise_all['Status'] == 'PAGO'].copy()

    # Define KPIs financeiros
    if df_analise_pago.empty:
        st.warning("Nenhuma venda 'PAGA' encontrada com a combinação de filtros selecionada no período. As métricas financeiras (Faturamento, Lucro, Ticket Médio) e gráficos de evolução estão zeradas.")
        total_faturamento = 0.0
        total_custo = 0.0
        lucro_bruto = 0.0
        total_vendas_paid_for_ticket = 0
    else:
        # --- KPIs PRINCIPAIS FINANCEIROS ---
        total_faturamento = df_analise_pago['Preço Venda Total'].sum()
        total_custo = df_analise_pago['Preço Custo Total'].sum() # Custo é registrado como negativo
        lucro_bruto = total_faturamento + total_custo # O custo é negativo, então soma
        # Contagem de vendas ÚNICAS que contribuíram com valor pago
        total_vendas_paid_for_ticket = df_analise_pago['ID_Venda'].nunique()


    # --- KPIs PRINCIPAIS ---
    st.markdown("---")
    st.subheader("Indicadores Chave de Desempenho (KPIs)")
    
    col_kpi_1, col_kpi_2, col_kpi_3, col_kpi_4 = st.columns(4)
    col_kpi_1.metric("Faturamento Bruto", f"R$ {total_faturamento:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))
    col_kpi_2.metric("Lucro Bruto", f"R$ {lucro_bruto:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))
    # TICKET MÉDIO: Faturamento Pago / Vendas Únicas Pagas
    col_kpi_3.metric("Ticket Médio", f"R$ {(total_faturamento / total_vendas_paid_for_ticket) if total_vendas_paid_for_ticket > 0 else 0:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))
    # TOTAL DE VENDAS: Vendas Únicas (Pagas ou A Receber)
    col_kpi_4.metric("Total de Vendas", total_vendas) 

    st.markdown("---")

    # Se não houver dados pagos, não gera os gráficos de análise de produto e evolução
    if df_analise_pago.empty:
        return

    # A partir daqui, os gráficos e análises usam df_analise_pago (apenas vendas realizadas)
    df_analise = df_analise_pago

    # --- ANÁLISES POR PRODUTO/CATEGORIA ---
    st.subheader("Análise de Desempenho por Item/Grupo")

    tab_prod, tab_cat, tab_col = st.tabs(["Produtos", "Categorias", "Coleções"])

    # --- TAB PRODUTOS: Top 50 Melhores e Piores (por Lucro Bruto) ---
    with tab_prod:
        
        # 1. Agrega por Produto (usando df_analise final)
        df_prod_ranking = df_analise.groupby('Produto').agg(
            Faturamento=('Preço Venda Total', 'sum'),
            Lucro=('Lucro Bruto', 'sum'),
            ItensVendidos=('Quantidade', lambda x: -x.sum()) # Quantidade é negativa, então inverte
        ).reset_index()
        
        # 2. Top 50 Melhores (Maior Lucro Bruto)
        df_melhores = df_prod_ranking.sort_values(by='Lucro', ascending=False).head(50).copy()
        
        # 3. Top 50 Piores (Menor Lucro Bruto/Maior Prejuízo)
        df_piores = df_prod_ranking.sort_values(by='Lucro', ascending=True).head(50).copy()

        # Gráfico para Top 10 Melhores (visualização rápida)
        st.markdown("##### 🏆 Top 10 Produtos por Lucro Bruto (Visualização Rápida)")
        
        df_top_10_visual = df_melhores.head(10)

        fig_prod_melhores = px.bar(
            df_top_10_visual.sort_values(by='Lucro', ascending=True),
            y='Produto',
            x='Lucro',
            color='Faturamento', # Usa faturamento como cor de contexto
            orientation='h',
            title='Produtos Mais Lucrativos (Lucro Bruto)',
            color_continuous_scale=px.colors.sequential.Plasma,
            labels={'Produto': '', 'Lucro': 'Lucro Bruto (R$)', 'Faturamento': 'Faturamento (R$)'}
        )
        fig_prod_melhores.update_layout(xaxis_tickformat='$,.2f', template="plotly_white")
        st.plotly_chart(fig_prod_melhores, use_container_width=True)
        
        st.markdown("---")

        # Tabelas para Top 50 Melhores e Piores
        col_tabela_1, col_tabela_2 = st.columns(2)
        
        with col_tabela_1:
            st.markdown("##### 🥇 TOP 50 Melhores Produtos (Maior Lucro Bruto)")
            
            # Formatação para exibição na tabela
            df_melhores_display = df_melhores.copy()
            df_melhores_display['Faturamento'] = df_melhores_display['Faturamento'].map('R$ {:,.2f}'.format).str.replace(',', 'X').str.replace('.', ',').str.replace('X', '.')
            df_melhores_display['Lucro'] = df_melhores_display['Lucro'].map('R$ {:,.2f}'.format).str.replace(',', 'X').str.replace('.', ',').str.replace('X', '.')

            st.dataframe(
                df_melhores_display[['Produto', 'Lucro', 'Faturamento', 'ItensVendidos']].rename(columns={'ItensVendidos': 'Itens'}),
                use_container_width=True,
                hide_index=True
            )
            
        with col_tabela_2:
            st.markdown("##### 📉 TOP 50 Piores Produtos (Menor Lucro Bruto)")
            
            # Formatação para exibição na tabela
            df_piores_display = df_piores.copy()
            df_piores_display['Faturamento'] = df_piores_display['Faturamento'].map('R$ {:,.2f}'.format).str.replace(',', 'X').str.replace('.', ',').str.replace('X', '.')
            df_piores_display['Lucro'] = df_piores_display['Lucro'].map('R$ {:,.2f}'.format).str.replace(',', 'X').str.replace('.', ',').str.replace('X', '.')
            
            st.dataframe(
                df_piores_display[['Produto', 'Lucro', 'Faturamento', 'ItensVendidos']].rename(columns={'ItensVendidos': 'Itens'}),
                use_container_width=True,
                hide_index=True
            )


    # --- TAB CATEGORIAS: Faturamento (Pizza) e Lucro (Barra) ---
    with tab_cat:
        # Usa df_analise final
        df_cat = df_analise.groupby('Categoria').agg(
            Faturamento=('Preço Venda Total', 'sum'),
            Lucro=('Lucro Bruto', 'sum')
        ).sort_values(by='Faturamento', ascending=False).reset_index()

        col_cat_1, col_cat_2 = st.columns(2)
        
        with col_cat_1:
            st.markdown("##### 🍕 Distribuição de Faturamento por Categoria")
            fig_cat_fat = px.pie(
                df_cat,
                values='Faturamento',
                names='Categoria',
                title='Proporção de Faturamento Total',
                color_discrete_sequence=px.colors.sequential.RdPu
            )
            fig_cat_fat.update_traces(textposition='inside', textinfo='percent+label', marker=dict(line=dict(color='#000000', width=1)))
            fig_cat_fat.update_layout(showlegend=False)
            st.plotly_chart(fig_cat_fat, use_container_width=True)

        with col_cat_2:
            st.markdown("##### 📈 Ranking de Lucro Bruto por Categoria")
            df_cat_lucro = df_cat.sort_values(by='Lucro', ascending=True)

            fig_cat_luc = px.bar(
                df_cat_lucro,
                y='Categoria',
                x='Lucro',
                orientation='h',
                title='Categoria com Maior Geração de Lucro',
                color_discrete_sequence=['#880e4f'], # Cor fixa do tema (Rosa Escuro)
                labels={'Categoria': '', 'Lucro': 'Lucro Bruto (R$)'}
            )
            fig_cat_luc.update_layout(xaxis_tickformat='$,.2f', template="plotly_white")
            st.plotly_chart(fig_cat_luc, use_container_width=True)


    # --- TAB COLEÇÕES: Faturamento (Colorido por Lucro) ---
    with tab_col:
        st.markdown("##### Faturamento por Coleção")
        
        # Usa df_analise final
        df_col = df_analise.groupby('Coleção').agg(
            Faturamento=('Preço Venda Total', 'sum'),
            Lucro=('Lucro Bruto', 'sum')
        ).sort_values(by='Faturamento', ascending=False).reset_index()
        
        fig_col = px.bar(
            df_col.sort_values(by='Faturamento', ascending=True),
            y='Coleção',
            x='Faturamento',
            color='Lucro',
            orientation='h',
            title='Faturamento por Coleção',
            color_continuous_scale=px.colors.sequential.RdPu,
            labels={'Coleção': '', 'Faturamento': 'Faturamento (R$)', 'Lucro': 'Lucro Bruto (R$)'}
        )
        fig_col.update_layout(xaxis_tickformat='$,.2f', template="plotly_white")
        st.plotly_chart(fig_col, use_container_width=True)


    st.markdown("---")
    
    # --- HISTOGRAMA DE FATURAMENTO (USANDO df_analise FINAL) ---
    st.subheader("Evolução do Faturamento")

    col_hist_1, col_hist_2 = st.columns([1, 3])
    with col_hist_1:
        granularity = st.radio(
            "Agrupar por:", 
            ["Semanal", "Mensal", "Anual"], 
            index=1,
            key='hist_granularity'
        )

    # --- Filtragem Inicial de Vendas Pagas (df_analise) ---
    # Assumindo que df_analise já foi criado na parte superior do seu dashboard
    # Exemplo de como df_analise deve ser criado (se já existir, ignore este trecho):
    # df_analise = df_movimento[
    #    (df_movimento['Tipo de Movimentação'] == 'SAÍDA') & 
    #    (df_movimento['Status'] == 'PAGO')
    # ].copy()
    # df_analise['Lucro Bruto'] = df_analise['Preço Venda Total'] + df_analise['Preço Custo Total']
    # df_analise['Data'] = pd.to_datetime(df_analise['Data']).dt.normalize()
    # --------------------------------------------------------

    if granularity == "Semanal":
        
        # 1. ENCONTRA AS SEMANAS
        # Identifica todas as semanas presentes nos dados para o selectbox
        df_analise['AnoSemana'] = df_analise['Data'].dt.strftime('%Y-W%W')
        opcoes_semanas = sorted(df_analise['AnoSemana'].unique(), reverse=True)
        
        # 2. SELECIONA A SEMANA
        # Adiciona um novo seletor para escolher a semana
        semana_selecionada = st.selectbox(
            "Selecione a Semana:", 
            opcoes_semanas,
            key='semana_selecionada'
        )
        
        # 3. FILTRA APENAS OS DADOS DA SEMANA ESCOLHIDA
        df_filtrado_semana = df_analise[df_analise['AnoSemana'] == semana_selecionada].copy()
        
        if df_filtrado_semana.empty:
            df_agrupado = pd.DataFrame(columns=['Data', 'Faturamento', 'Lucro', 'Periodo'])
            hover_name = 'Dia'
        else:
            # 4. AGRUPA DIA A DIA
            df_agrupado = df_filtrado_semana.set_index('Data').resample('D').agg(
                Faturamento=('Preço Venda Total', 'sum'),
                Lucro=('Lucro Bruto', 'sum')
            ).reset_index()

            # Remove dias sem vendas (Faturamento = 0) se desejar, ou mantém para ter o dia na linha do tempo.
            df_agrupado = df_agrupado[df_agrupado['Faturamento'] > 0] 
            
            # Cria o nome do dia para o eixo X
            df_agrupado['Periodo'] = df_agrupado['Data'].dt.strftime('%d/%m (%a)') 
            hover_name = 'Dia'

    elif granularity == "Mensal":
        # Mensal: 'ME'
        df_agrupado = df_analise.set_index('Data').resample('ME').agg(
            Faturamento=('Preço Venda Total', 'sum'),
            Lucro=('Lucro Bruto', 'sum')
        ).reset_index()
        df_agrupado['Periodo'] = df_agrupado['Data'].dt.strftime('%b/%Y')
        hover_name = 'Mês'
    else: # Anual
        # Anual: 'YE'
        df_agrupado = df_analise.set_index('Data').resample('YE').agg(
            Faturamento=('Preço Venda Total', 'sum'),
            Lucro=('Lucro Bruto', 'sum')
        ).reset_index()
        df_agrupado['Periodo'] = df_agrupado['Data'].dt.year.astype(str)
        hover_name = 'Ano'

    with col_hist_2:
        if df_agrupado.empty:
            st.info(f"Sem dados de vendas para o período selecionado ({granularity}).")
        else:
            fig_hist = px.bar(
                df_agrupado, 
                x='Periodo', 
                y='Faturamento', 
                color='Lucro', 
                title=f'Evolução do Faturamento Bruto por {hover_name}',
                labels={'Periodo': hover_name, 'Faturamento': 'Faturamento (R$)', 'Lucro': 'Lucro Bruto (R$)'},
                template="plotly_white",
                color_continuous_scale=px.colors.sequential.Agsunset, 
                text='Faturamento' 
            )
            
            # Garante a ordem correta no eixo X
            fig_hist.update_layout(
                xaxis={'categoryorder':'array', 'categoryarray':df_agrupado['Periodo'].tolist()},
                yaxis={'tickprefix': 'R$ ', 'separatethousands': True},
                coloraxis_colorbar=dict(
                    title="Lucro (R$)",
                    tickprefix="R$ "
                )
            )
            
            fig_hist.update_traces(
                texttemplate='R$%{text:,.2f}', 
                textposition='outside'
            )
            
            st.plotly_chart(fig_hist, use_container_width=True)

def registrar_entrada():
    """Interface para registro de uma nova entrada (compra/reposição de estoque) com data customizada."""
    st.title("Registrar Nova Entrada (Compra/Reposição)")
    
    if "carrinho_entrada" not in st.session_state:
        st.session_state["carrinho_entrada"] = []

    df_produtos_validos = carregar_produtos()

    if df_produtos_validos.empty:
        st.warning("Não foi possível carregar a base de produtos válidos.")
        return
    
    # ⚠️ NOVO CAMPO: Data de Registro (Personalizada)
    data_registro_entrada = st.date_input(
        "Data da Entrada (Compra)",
        value=date.today(), # Padrão é hoje
        max_value=date.today(), # Não permite datas futuras
        key="data_registro_entrada_custom"
    )
    st.markdown("---")
    
    # Prepara opções do seletor: COD - Nome do Produto
    opcoes_produtos = (df_produtos_validos['COD'] + " - " + df_produtos_validos['Marca'] + " "+df_produtos_validos['Produto']).tolist()
    opcoes_produtos.insert(0, "Selecione um produto...")
    
    st.subheader("1. Adicionar Item para Entrada")
    
    col1, col2, col3, col4 = st.columns([3, 1.5, 1.5, 1.5])
    
    with col1:
        produto_selecionado_add = st.selectbox(
            "Produto", 
            opcoes_produtos,
            key='selectbox_produto_add_entrada' 
        )

    produto_info = None
    cod_selecionado = None
    preco_custo_unitario_default = 0.0
    
    if produto_selecionado_add != "Selecione um produto...":
        cod_selecionado = produto_selecionado_add.split(" - ")[0].strip()
        produto_info_df = df_produtos_validos[df_produtos_validos['COD'] == cod_selecionado]
        if not produto_info_df.empty:
            produto_info = produto_info_df.iloc[0]
            try:
                preco_custo_default = pd.to_numeric(produto_info.get('Preço Custo', 0), errors='coerce')
                preco_custo_unitario_default = float(preco_custo_default) if pd.notna(preco_custo_default) else 0.0
            except Exception:
                preco_custo_unitario_default = 0.0

    with col2:
        qtd_add = st.number_input("Quantidade da Entrada", min_value=1, value=1, step=1, key='input_quantidade_add_entrada')
    
    with col3:
        preco_custo_unitario_add = st.number_input(
            "Preço Custo Unit. (Pago)", 
            min_value=0.0, 
            value=preco_custo_unitario_default,
            format="%.2f",
            key='input_custo_add_entrada'
        )
    
    with col4:
        st.markdown("<br>", unsafe_allow_html=True) 
        if st.button("Adicionar Item", key='btn_add_carrinho_entrada'):
            if produto_selecionado_add == "Selecione um produto...":
                st.error("Selecione um produto.")
            elif cod_selecionado is None:
                st.error("Erro ao carregar informações do produto.")
            elif qtd_add <= 0:
                st.error("A quantidade deve ser maior que zero.")
            elif preco_custo_unitario_add <= 0:
                st.warning("O Preço Custo Unitário não pode ser zero. Por favor, corrija.")
            else:
                # Nome do produto (com marca, sem hífen)
                if produto_info is not None:
                    marca = produto_info.get('Marca', '').strip()
                    nome_base = produto_info.get('Produto', '').strip()
                    nome_prod = f"{marca} {nome_base}" if marca else nome_base
                else:
                    nome_prod = produto_selecionado_add.split(" - ", 1)[1]
                
                item = {
                    "COD do Produto": cod_selecionado,
                    "Produto": nome_prod,
                    "Quantidade": qtd_add,
                    "Preço Custo Unitário": preco_custo_unitario_add,
                    "Preço Custo Total": qtd_add * preco_custo_unitario_add,
                    "Preço Venda Unitário": produto_info.get('Preço Venda', 0.0) if produto_info is not None else 0.0,
                    "Preço Venda Total": (qtd_add * produto_info.get('Preço Venda', 0.0)) if produto_info is not None else 0.0,
                }
                st.session_state["carrinho_entrada"].append(item)
                st.success(f"{qtd_add}x {item['Produto']} adicionado(s) à entrada.")
                time.sleep(1)
                st.rerun() 
                return

    st.markdown("---")
    
    st.subheader("2. Resumo da Entrada e Pagamento")

    if not st.session_state.get("carrinho_entrada"):
        st.info("O carrinho de entrada está vazio. Adicione um produto para continuar.")
        return

    df_carrinho = pd.DataFrame(st.session_state["carrinho_entrada"])
    
    df_display = df_carrinho[[
        "Produto", "Quantidade", "Preço Custo Unitário", "Preço Custo Total"
    ]].copy()
    for col in ["Preço Custo Unitário", "Preço Custo Total"]:
        df_display[col] = df_display[col].apply(lambda x: f"R$ {x:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))

    st.dataframe(df_display, use_container_width=True)
    
    if st.button("Limpar Carrinho de Entrada", key='btn_limpar_carrinho_entrada'):
        st.session_state["carrinho_entrada"] = []
        st.rerun()
        return

    total_custo = df_carrinho["Preço Custo Total"].sum()
    st.markdown(f"#### **Custo Total da Compra: R$ {total_custo:,.2f}**".replace(",", "X").replace(".", ",").replace("X", "."))

    # Pagamento (Foco no custo/saída financeira)
    col_status, col_meio = st.columns(2)
    with col_status:
        status_pagamento = st.selectbox(
            "Status da Compra",
            ["Pago (Saída Financeira)", "A Pagar (Conta)"],
            key="status_pagamento_entrada_final"
        )
    
    with col_meio:
        if status_pagamento == "Pago (Saída Financeira)":
            meio_pagamento = st.radio(
                "Meio de Pagamento",
                ["Pix", "Cartão", "Dinheiro"],
                key="meio_pagamento_entrada_radio",
                horizontal=True
            )
            data_prevista_vencimento = data_registro_entrada.strftime('%Y-%m-%d 23:59:59') # Assume pagamento no mesmo dia customizado
        else:
            meio_pagamento = "Conta a Pagar"
            data_prevista_vencimento = st.date_input(
                "Data Prevista de Pagamento",
                datetime.today() + DateOffset(days=7),
                key="data_prevista_compra_final"
            ).strftime('%Y-%m-%d 00:00:00')

    with st.form("form_finalizar_entrada"):
        fornecedor_final = st.text_input("Fornecedor (Opcional)", value="", key='input_fornecedor_final_form')
        Observacoes_final = st.text_area("Observações (opcional)", value="", key="Observacoes_key_final_form_entrada")
        submitted = st.form_submit_button("Finalizar Entrada e Registrar Movimento")
        
        if submitted:
            if total_custo <= 0: 
                st.error("O Custo Total da Entrada deve ser maior que zero.")
            else:
                df = carregar_dados()
                
                id_movimento = f"E-{uuid.uuid4().hex[:6].upper()}"
                status_registro = "PAGO" if status_pagamento == "Pago (Saída Financeira)" else "A PAGAR"
                
                # ⚠️ USANDO A DATA PERSONALIZADA AQUI:
                # Se a data for retroativa, registra no início do dia
                # Se a data for hoje, registra a hora atual.
                data_registro_final = (
                    datetime.now() if data_registro_entrada == date.today() 
                    else datetime.combine(data_registro_entrada, datetime.min.time())
                )
                
                registros = []
                
                # 1. Registro de ENTRADA de Estoque (Quantidade Positiva)
                for item in st.session_state["carrinho_entrada"]:
                    registros.append({
                        "Data": data_registro_final, # ⚠️ DATA PERSONALIZADA
                        "COD do Produto": item["COD do Produto"],
                        "Produto": item["Produto"],
                        "Cliente": fornecedor_final, 
                        "Tipo de Movimentação": "ENTRADA",
                        "Quantidade": item["Quantidade"],
                        "Preço Custo Total": item["Preço Custo Total"], 
                        "Preço Venda Total": item["Preço Venda Total"],
                        "Observações": Observacoes_final.strip(), 
                        "Status": status_registro,
                        "Data Prevista": data_prevista_vencimento, 
                        "Tipo de Pagamento": meio_pagamento, 
                        "ID_Venda": id_movimento,
                    })

                df_novo = pd.concat([df, pd.DataFrame(registros)], ignore_index=True)
                
                save_successful = salvar_dados(df_novo)  
                
                if save_successful: 
                    st.session_state["carrinho_entrada"] = [] 
                    st.success(f"Entrada {id_movimento} registrada com sucesso!")
                    time.sleep(2) 
                    st.rerun() 
                else:
                    st.error("Falha ao registrar entrada. Verifique as mensagens de erro acima.")
# ===== PÁGINAS (PAINEIS) DO main2 copy 2.py (Adaptadas) =====

def page_dashboard():
# ... (Restante do código mantido)
    st.markdown("<div class='big-title'>DASHBOARD</div>", unsafe_allow_html=True)
    st.markdown("<div class='subtitle'>Visão geral | Filtros rápidos | KPIs</div>", unsafe_allow_html=True)
    
    # Carrega os dados necessários para o dashboard
    df_movimento = carregar_dados()
    df_produtos = carregar_produtos()
    
    if df_movimento.empty or df_produtos.empty:
        st.warning("Não foi possível carregar os dados de Movimento ou Produtos para o Dashboard. Verifique seu arquivo BD_Loja.xlsx.")
        return
        
    page_dashboard_logic(df_movimento, df_produtos)

def page_sales():
# ... (Restante do código mantido)
    """Painel de Vendas utilizando a lógica do main.py (main1)."""
    st.markdown("<div class='big-title'>VENDAS & RECEBIMENTOS</div>", unsafe_allow_html=True)
    st.markdown("<div class='subtitle'>Ações de Vendas e Pagamento</div>", unsafe_allow_html=True)
    
    acao_vendas = st.selectbox("Ações de Vendas", [
        "Registrar Venda", 
        "Atualizar Pagamento",
        "Registrar Entrada",
        "Ver Saldo de Clientes",
        "Ver Todas as Vendas",
    ], key="vendas_main1_actions")

    if acao_vendas == "Registrar Venda":
        registrar_venda() 
    elif acao_vendas == "Atualizar Pagamento":
        atualizar_recebimento() 
    elif acao_vendas == "Ver Saldo de Clientes":
        mostrar_saldos() 
    elif acao_vendas == "Ver Todas as Vendas":
        mostrar_todas_vendas() 
    elif acao_vendas == "Registrar Entrada":
        registrar_entrada() 

def page_products():
    st.markdown("<div class='big-title'>PRODUTOS & ESTOQUE</div>", unsafe_allow_html=True)
    st.markdown("<div class='subtitle'>Listagem, Controle e Atualização de Preços</div>", unsafe_allow_html=True)
    
    # Carregar Dados
    df_produtos_base = carregar_produtos() 
    df_estoque = calcular_estoque()

    if df_produtos_base.empty:
        st.error("Não foi possível carregar a base de produtos.")
        return
    
    # Seletor de Ações
    acao_produtos = st.selectbox("Ações de Produtos", [
        "Ver Lista e Estoque",
        "Atualizar Preço de Produto",
    ], key="produtos_actions")

    if acao_produtos == "Ver Lista e Estoque":
        page_products_list(df_produtos_base, df_estoque)
    elif acao_produtos == "Atualizar Preço de Produto":
        atualizar_produto(df_produtos_base)

def page_clients():
    st.markdown("<div class='big-title'>CLIENTES & CONTATOS</div>", unsafe_allow_html=True)
    st.markdown("<div class='subtitle'>Cadastro | Edição | Lista de Contatos</div>", unsafe_allow_html=True)

    # Carrega os dados dos clientes
    df_clientes = carregar_clientes()
    df_movimento = carregar_dados() # Carrega para calcular "Total Gasto" e "Última Compra"

    # Seletor de Ações
    acao_clientes = st.selectbox(
        "Ações de Clientes",
        ["Listar Clientes", "Cadastrar Novo Cliente", "Editar Cliente", "Excluir Cliente"],
        key="clientes_actions"
    )
    
    st.markdown("---")

    if acao_clientes == "Listar Clientes":
        st.subheader("Lista de Clientes Cadastrados")
        
        # 1. CALCULAR MÉTRICAS DINÂMICAS
        # Filtra apenas vendas com valor
        df_vendas = df_movimento[
            (df_movimento['Tipo de Movimentação'] == 'SAÍDA') & 
            (df_movimento['Preço Venda Total'] > 0)
        ].copy()
        
        df_display = df_clientes.copy()
        
        if not df_vendas.empty:
            # Calcula o Total Gasto
            df_gasto_total = df_vendas.groupby('Cliente')['Preço Venda Total'].sum().reset_index()
            df_gasto_total.rename(columns={'Preço Venda Total': 'Total Gasto'}, inplace=True)
            
            # Calcula a Última Compra
            df_ultima_compra = df_vendas.groupby('Cliente')['Data'].max().reset_index()
            df_ultima_compra.rename(columns={'Data': 'Última Compra'}, inplace=True)
            
            # Merge com a lista de clientes
            df_display = pd.merge(df_display, df_gasto_total, left_on='Nome', right_on='Cliente', how='left')
            df_display = pd.merge(df_display, df_ultima_compra, left_on='Nome', right_on='Cliente', how='left')
            
            # Limpa colunas extras do merge
            df_display.drop(columns=['Cliente_x', 'Cliente_y'], errors='ignore', inplace=True)
            df_display.drop(columns=['Cliente'], errors='ignore', inplace=True)
        else:
            df_display['Total Gasto'] = 0.0
            df_display['Última Compra'] = pd.NaT

        # Formatação das novas colunas
        df_display['Total Gasto'] = df_display['Total Gasto'].fillna(0)
        df_display['Última Compra'] = pd.to_datetime(df_display['Última Compra'], errors='coerce')
        
        df_display['Total Gasto (R$)'] = df_display['Total Gasto'].map('{:,.2f}'.format).str.replace(',', 'X').str.replace('.', ',').str.replace('X', '.')
        df_display['Última Compra (Data)'] = df_display['Última Compra'].dt.strftime('%d/%m/%Y').fillna('N/A')

        
        termo_busca = st.text_input("Buscar Cliente (Nome, Telefone, Email, Endereço)", key="busca_cliente")
        
        df_filtrado = df_display
        if termo_busca:
            termo_busca_lower = termo_busca.lower()
            colunas_busca_cliente = ['Nome', 'Telefone', 'Email', 'Endereço', 'Observações']
            
            busca_mask = df_filtrado[colunas_busca_cliente].astype(str).apply(
                lambda x: x.str.lower().str.contains(termo_busca_lower, na=False)
            ).any(axis=1)
            df_filtrado = df_filtrado[busca_mask]
            
        # Define a ordem das colunas para exibição
        colunas_exibir_final = [
            'Nome', 'Telefone', 'Total Gasto (R$)', 'Última Compra (Data)', 
            'Email', 'Endereço', 'Observações', 'ID_Cliente'
        ]
        colunas_existentes = [col for col in colunas_exibir_final if col in df_filtrado.columns]
        
        st.dataframe(
            df_filtrado[colunas_existentes].sort_values(by="Nome"), 
            use_container_width=True, 
            hide_index=True
        )
        st.info("Nota: 'Total Gasto' e 'Última Compra' são calculados com base nas vendas registradas. O cálculo depende da exatidão do nome do cliente no momento da venda.")

    elif acao_clientes == "Cadastrar Novo Cliente":
        st.subheader("Cadastrar Novo Cliente")
        with st.form("form_novo_cliente", clear_on_submit=True):
            nome = st.text_input("Nome *", key="novo_cliente_nome")
            telefone = st.text_input("Telefone", key="novo_cliente_tel")
            email = st.text_input("Email", key="novo_cliente_email")
            endereco = st.text_area("Endereço", key="novo_cliente_end")
            obs = st.text_area("Observações", key="novo_cliente_obs")
            
            submitted = st.form_submit_button("Cadastrar Cliente")
            
            if submitted:
                if not nome:
                    st.error("O campo 'Nome' é obrigatório.")
                else:
                    # Checa duplicidade de nome
                    if nome.strip().lower() in df_clientes['Nome'].str.strip().str.lower().values:
                        st.error(f"Cliente com o nome '{nome}' já existe.")
                    else:
                        novo_id = f"C-{uuid.uuid4().hex[:6].upper()}"
                        novo_cliente_data = {
                            "ID_Cliente": novo_id,
                            "Nome": nome.strip(),
                            "Telefone": telefone.strip(),
                            "Email": email.strip(),
                            "Endereço": endereco.strip(),
                            "Observações": obs.strip()
                        }
                        
                        df_novo = pd.concat([df_clientes, pd.DataFrame([novo_cliente_data])], ignore_index=True)
                        
                        if salvar_clientes(df_novo):
                            st.success(f"Cliente '{nome}' (ID: {novo_id}) cadastrado com sucesso!")
                            time.sleep(1)
                            st.rerun()
                        else:
                            st.error("Falha ao salvar o novo cliente.")

    elif acao_clientes == "Editar Cliente":
        st.subheader("Editar Cliente Existente")
        if df_clientes.empty:
            st.warning("Nenhum cliente cadastrado para editar.")
            return

        opcoes_clientes = ["Selecione um cliente..."] + df_clientes.sort_values(by="Nome")['Nome'].tolist()
        cliente_nome_selecionado = st.selectbox("Selecione o Cliente para Editar", opcoes_clientes, key="editar_cliente_select")
        
        if cliente_nome_selecionado != "Selecione um cliente...":
            
            # Pega os dados do cliente (pode haver nomes duplicados, pega o primeiro)
            cliente_data_series = df_clientes[df_clientes['Nome'] == cliente_nome_selecionado].iloc[0]
            cliente_data = cliente_data_series.to_dict()
            cliente_id_seguro = cliente_data['ID_Cliente']
            
            with st.form("form_editar_cliente"):
                st.text(f"Editando Cliente (ID: {cliente_id_seguro})")
                
                nome_edit = st.text_input("Nome *", value=cliente_data['Nome'], key="edit_cliente_nome")
                telefone_edit = st.text_input("Telefone", value=cliente_data['Telefone'], key="edit_cliente_tel")
                email_edit = st.text_input("Email", value=cliente_data['Email'], key="edit_cliente_email")
                endereco_edit = st.text_area("Endereço", value=cliente_data['Endereço'], key="edit_cliente_end")
                obs_edit = st.text_area("Observações", value=cliente_data['Observações'], key="edit_cliente_obs")
                
                submitted_edit = st.form_submit_button("Atualizar Cliente")
                
                if submitted_edit:
                    if not nome_edit:
                        st.error("O campo 'Nome' é obrigatório.")
                    else:
                        # Pega o índice do cliente original (pelo ID, que é a chave segura)
                        idx_original_list = df_clientes.index[df_clientes['ID_Cliente'] == cliente_id_seguro].tolist()
                        
                        if not idx_original_list:
                             st.error("Erro: ID do cliente não encontrado. Não foi possível salvar.")
                             return
                        
                        idx_original = idx_original_list[0]
                        
                        # Atualiza os dados no DataFrame
                        df_clientes.loc[idx_original, 'Nome'] = nome_edit.strip()
                        df_clientes.loc[idx_original, 'Telefone'] = telefone_edit.strip()
                        df_clientes.loc[idx_original, 'Email'] = email_edit.strip()
                        df_clientes.loc[idx_original, 'Endereço'] = endereco_edit.strip()
                        df_clientes.loc[idx_original, 'Observações'] = obs_edit.strip()
                        
                        if salvar_clientes(df_clientes):
                            st.success(f"Dados do cliente '{nome_edit}' atualizados com sucesso!")
                            
                            if nome_edit.strip() != cliente_data['Nome'].strip():
                                st.warning(f"Atenção: O nome do cliente foi alterado de '{cliente_data['Nome']}' para '{nome_edit}'. Registros de vendas *anteriores* em 'Histórico de Vendas' e 'Saldos' ainda podem exibir o nome antigo.")
                            
                            time.sleep(2)
                            st.rerun()
                        else:
                            st.error("Falha ao salvar a atualização.")

    elif acao_clientes == "Excluir Cliente":
        st.subheader("Excluir Cliente")
        st.warning("🚨 Atenção: Esta ação é irreversível.", icon="⚠️")
        if df_clientes.empty:
            st.warning("Nenhum cliente cadastrado para excluir.")
            return
            
        opcoes_clientes_del = ["Selecione um cliente para excluir..."] + df_clientes.sort_values(by="Nome")['Nome'].tolist()
        cliente_nome_del = st.selectbox("Cliente", opcoes_clientes_del, key="excluir_cliente_select")
        
        if cliente_nome_del != "Selecione um cliente para excluir...":
            cliente_data_del = df_clientes[df_clientes['Nome'] == cliente_nome_del].iloc[0]
            st.markdown(f"Você está prestes a excluir **{cliente_data_del['Nome']}** (ID: {cliente_data_del['ID_Cliente']}).")
            
            # Checagem de segurança: O cliente tem vendas?
            # Compara nomes sem case e espaços para segurança
            vendas_cliente = df_movimento[
                df_movimento['Cliente'].str.strip().str.lower() == cliente_data_del['Nome'].strip().lower()
            ]
            
            if not vendas_cliente.empty:
                st.error(f"Não é possível excluir '{cliente_data_del['Nome']}'.")
                st.info(f"Este cliente já possui {len(vendas_cliente)} registro(s) de movimentação (vendas). Para preservar o histórico financeiro, clientes com movimentações não podem ser excluídos.")
            else:
                if st.button(f"Confirmar Exclusão Definitiva de {cliente_data_del['Nome']}", key="btn_confirmar_excluir_cliente"):
                    
                    df_filtrado = df_clientes[df_clientes['ID_Cliente'] != cliente_data_del['ID_Cliente']]
                    
                    if salvar_clientes(df_filtrado):
                        st.success(f"Cliente '{cliente_data_del['Nome']}' excluído com sucesso.")
                        time.sleep(2)
                        st.rerun()
                    else:
                        st.error("Falha ao salvar a exclusão.")

                        
def page_reports():
    st.markdown("<div class='big-title'>RELATÓRIOS E ANÁLISES FINANCEIRAS</div>", unsafe_allow_html=True)
    st.markdown("<div class='subtitle'>Acompanhamento de Dívidas, Lucros e Estoque</div>", unsafe_allow_html=True)

    df_movimento = carregar_dados()
    df_produtos = carregar_produtos()
    
    if df_movimento.empty:
        st.warning("Não há dados de movimentação para gerar relatórios.")
        return

    report_type = st.selectbox(
        "Selecione o Relatório",
        [ "Análise de Lucros e Margem", "Custo Total do Estoque", "Movimentação de Caixa/Estoque (Entrada x Saída)","Análises de Desempenho e Clientes", "Devedores (Contas a Receber)"],
        key="report_selector"
    )
    
    st.markdown("---")

    if report_type == "Devedores (Contas a Receber)":
        st.subheader("Relatório de Contas a Receber (Devedores)")
        
        # 1. Devedores (A RECEBER)
        df_devedores = df_movimento[
            df_movimento['Status'] == 'A RECEBER'
        ].copy()

        if df_devedores.empty:
            st.success("🎉 Não há contas a receber ou clientes devedores no momento.")
            return

        # Calcular o saldo devedor por cliente
        df_saldo_devedor = df_devedores.groupby('Cliente')['Preço Venda Total'].sum().reset_index()
        df_saldo_devedor.rename(columns={'Preço Venda Total': 'Dívida Total (R$)'}, inplace=True)
        
        # Formatar a dívida total
        df_saldo_devedor['Dívida Total (R$)'] = df_saldo_devedor['Dívida Total (R$)'].apply(lambda x: f"R$ {x:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))
        
        st.markdown("#### Saldo Devedor Consolidado por Cliente")
        st.dataframe(df_saldo_devedor.sort_values(by='Dívida Total (R$)', ascending=False), hide_index=True, use_container_width=True)

        # Detalhe das parcelas/vendas
        st.markdown("#### Detalhamento das Contas a Receber")
        
        # Limpar e ordenar as colunas de detalhe
        df_detalhe = df_devedores[[
            'Data Prevista', 'Cliente', 'Produto', 'Preço Venda Total', 'Observações', 'ID_Venda'
        ]].copy()
        
        df_detalhe['Data Prevista'] = pd.to_datetime(df_detalhe['Data Prevista'], errors='coerce').dt.strftime('%d/%m/%Y')
        df_detalhe['Preço Venda Total'] = df_detalhe['Preço Venda Total'].apply(lambda x: f"R$ {x:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))
        df_detalhe.rename(columns={
            'Preço Venda Total': 'Valor a Receber (R$)',
            'Data Prevista': 'Vencimento',
            'Observações': 'Detalhes'
        }, inplace=True)

        st.dataframe(
            df_detalhe.sort_values(by='Vencimento', ascending=True), 
            hide_index=True, 
            use_container_width=True
        )
        
    elif report_type == "Análise de Lucros e Margem":
        st.subheader("Relatório de Lucros e Margem")
        
        # 2. Lucros
        df_saidas = df_movimento[df_movimento['Tipo de Movimentação'] == 'SAÍDA'].copy()

        if df_saidas.empty:
            st.info("Não há vendas (saídas) registradas para calcular lucros.")
            return
            
        df_saidas['Preço Venda Total'] = pd.to_numeric(df_saidas['Preço Venda Total'], errors='coerce')
        df_saidas['Preço Custo Total'] = pd.to_numeric(df_saidas['Preço Custo Total'], errors='coerce')
        
        # A. Lucro Bruto Total (Inclui A RECEBER)
        receita_bruta = df_saidas['Preço Venda Total'].sum()
        custo_bruto = abs(df_saidas['Preço Custo Total'].sum()) 
        lucro_bruto = receita_bruta - custo_bruto
        
        # B. Lucro Realizado (Apenas PAGO)
        df_pagas = df_saidas[df_saidas['Status'] == 'PAGO']
        receita_paga = df_pagas['Preço Venda Total'].sum()
        custo_pago = abs(df_pagas['Preço Custo Total'].sum()) 
        lucro_realizado = receita_paga - custo_pago
        
        # C. Margem (usando Lucro Bruto)
        margem_bruta = (lucro_bruto / receita_bruta) * 100 if receita_bruta > 0 else 0

        col_l1, col_l2 = st.columns(2)
        with col_l1:
            st.metric(
                label="💰 Lucro Bruto Total (Vendas Realizadas + A Receber)", 
                value=f"R$ {lucro_bruto:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            )
            st.metric(
                label="📈 Margem Bruta Total", 
                value=f"{margem_bruta:,.2f}%".replace(",", "X").replace(".", ",").replace("X", ".")
            )
        with col_l2:
            st.metric(
                label="✅ Lucro Realizado (Apenas Vendas Pagas)", 
                value=f"R$ {lucro_realizado:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            )

            
        st.markdown("#### Histórico Mensal de Lucros (Bruto)")
        
        # Agrupamento Mensal
        df_saidas['Mês/Ano'] = pd.to_datetime(df_saidas['Data']).dt.to_period('M')
        
        df_mensal = df_saidas.groupby('Mês/Ano').agg(
            Receita=('Preço Venda Total', 'sum'),
            Custo=('Preço Custo Total', lambda x: abs(x.sum()))
        ).reset_index()
        
        df_mensal['Lucro Bruto'] = df_mensal['Receita'] - df_mensal['Custo']
        df_mensal['Mês/Ano'] = df_mensal['Mês/Ano'].astype(str)
        
        # Gráfico (usando Plotly)
        fig = px.bar(
            df_mensal, 
            x='Mês/Ano', 
            y='Lucro Bruto', 
            text='Lucro Bruto',
            title='Lucro Bruto por Mês',
            labels={'Lucro Bruto': 'Lucro (R$)', 'Mês/Ano': 'Mês'},
            color_discrete_sequence=px.colors.qualitative.Plotly
        )
        fig.update_traces(texttemplate='R$%{text:,.2f}', textposition='outside')
        fig.update_layout(uniformtext_minsize=8, uniformtext_mode='hide')
        st.plotly_chart(fig, use_container_width=True)


    elif report_type == "Custo Total do Estoque":
        st.subheader("Cálculo do Custo Total do Estoque Atual")
        
        # 3. Custo de Estoque
        
        # 3.1 Calcular Estoque Atual (Quantidade)
        df_estoque_mov = df_movimento.groupby('COD do Produto')['Quantidade'].sum().reset_index()
        df_estoque_mov.rename(columns={'Quantidade': 'Estoque Atual'}, inplace=True)
        
        df_estoque_mov = df_estoque_mov[df_estoque_mov['Estoque Atual'] > 0]

        if df_estoque_mov.empty:
            st.info("Não há produtos em estoque com base nas movimentações.")
            return
            
        # 3.2 Determinar o Custo Unitário de Avaliação
        df_estoque_custo = pd.merge(
            df_estoque_mov, 
            df_produtos[['COD', 'Marca', 'Produto', 'Preço Custo']], 
            left_on='COD do Produto', 
            right_on='COD', 
            how='left'
        )
        
        df_estoque_custo['Nome Completo'] = df_estoque_custo['Marca'] + " " + df_estoque_custo['Produto']
        df_estoque_custo['Preço Custo'] = pd.to_numeric(df_estoque_custo['Preço Custo'], errors='coerce').fillna(0)
        
        # 3.3 Calcular o Custo Total
        df_estoque_custo['Custo Total do Item (R$)'] = df_estoque_custo['Estoque Atual'] * df_estoque_custo['Preço Custo']
        
        custo_total_estoque = df_estoque_custo['Custo Total do Item (R$)'].sum()
        
        st.metric(
            label="📦 Custo Total de Avaliação do Estoque",
            value=f"R$ {custo_total_estoque:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        )
        
        st.markdown("#### Detalhamento do Custo por Produto em Estoque")
        
        df_display_estoque = df_estoque_custo[[
            'Nome Completo', 'Estoque Atual', 'Preço Custo', 'Custo Total do Item (R$)'
        ]].copy()
        
        df_display_estoque.rename(columns={'Preço Custo': 'Custo Unitário (R$)'}, inplace=True)

        for col in ['Custo Unitário (R$)', 'Custo Total do Item (R$)']:
            df_display_estoque[col] = df_display_estoque[col].apply(lambda x: f"R$ {x:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))

        st.dataframe(
            df_display_estoque.sort_values(by='Custo Total do Item (R$)', ascending=False),
            hide_index=True,
            use_container_width=True
        )
        st.info("O 'Custo Unitário' é o valor de custo cadastrado na página de Produtos, usado como base para a avaliação.")


    elif report_type == "Movimentação de Caixa/Estoque (Entrada x Saída)":
        st.subheader("Relatório Comparativo de Movimentação (Entrada x Saída)")
        
        # 4. Entrada x Movimento
        # Cria a coluna Mês/Ano para agrupamento
        df_movimento['Mês/Ano'] = pd.to_datetime(df_movimento['Data']).dt.to_period('M').astype(str)
        
        # CORREÇÃO CRÍTICA NO AGRUPAMENTO:
        # 1. Usamos a função 'sum' simples para Itens (Quantidade), garantindo a soma correta.
        # 2. Aplicamos .abs() nas colunas Custo e Itens APÓS o agrupamento para garantir valores positivos,
        #    sem interferir na agregação do Pandas.
        df_agrupado = df_movimento.groupby(['Mês/Ano', 'Tipo de Movimentação']).agg(
            Valor=('Preço Venda Total', 'sum'),
            Custo=('Preço Custo Total', 'sum'),
            Itens=('Quantidade', 'sum') # <-- CORRIGIDO: Agora usa a função de soma padrão
        ).reset_index()
        
        # Garantir que Custo e Itens sejam sempre positivos
        df_agrupado['Custo'] = df_agrupado['Custo'].abs()
        df_agrupado['Itens'] = df_agrupado['Itens'].abs()

        # Pivotar para ter ENTRADA e SAÍDA como colunas
        df_pivot = df_agrupado.pivot_table(
            index='Mês/Ano', 
            columns='Tipo de Movimentação', 
            values=['Custo', 'Itens', 'Valor'] # ORDEM ALFABÉTICA PADRÃO: Custo, Itens, Valor
        ).fillna(0).reset_index()
        
        # Renomear colunas para formato FLAT
        # A ordem padrão do Pandas é (Custo, Itens, Valor) x (ENTRADA, SAÍDA)
        df_pivot.columns = ['Mês/Ano', 
                            'Custo_ENTRADA', 'Custo_SAÍDA',
                            'Itens_ENTRADA', 'Itens_SAÍDA', 
                            'Valor_ENTRADA', 'Valor_SAÍDA', 
                        ]

        st.markdown("#### Comparação de Valor Financeiro por Mês")
        
        # Gráfico de Receita vs Custo (Total de Vendas)
        df_saida_agrupado = df_agrupado[df_agrupado['Tipo de Movimentação'] == 'SAÍDA'].copy()
        df_saida_agrupado.rename(columns={'Valor': 'Receita de Venda', 'Custo': 'Custo de Venda'}, inplace=True)

        # CORREÇÃO NO GRÁFICO: O Plotly Express cria 'variable' e não precisa de 'color' quando 'y' é lista.
        fig_valor = px.bar(
            df_saida_agrupado, 
            x='Mês/Ano', 
            y=['Receita de Venda', 'Custo de Venda'], 
            barmode='group',
            title='Receita de Venda vs Custo dos Itens Vendidos por Mês',
            labels={'value': 'Valor (R$)', 'variable': 'Métrica'},
        )
        fig_valor.for_each_trace(lambda t: t.update(name = t.name.replace('Receita de Venda', 'Receita').replace('Custo de Venda', 'Custo')))
        
        fig_valor.update_layout(
            legend_title_text='Métrica', 
            colorway=['#4caf50', '#f44336'] 
        )
        fig_valor.update_traces(hovertemplate='Mês: %{x}<br>Valor: R$%{y:,.2f}<extra></extra>')

        st.plotly_chart(fig_valor, use_container_width=True)


        st.markdown("#### Tabela Detalhada por Mês")
        
        # CORREÇÃO FINAL NA SELEÇÃO DE COLUNAS (resolve o KeyError anterior e remove a duplicata)
        df_display_mov = df_pivot[[
            'Mês/Ano', 
            'Itens_ENTRADA', 'Itens_SAÍDA', 
            'Custo_ENTRADA', 'Valor_ENTRADA', # Adicionado Valor_ENTRADA para clareza
            'Valor_SAÍDA', 
            'Custo_SAÍDA'
        ]].copy()
                
        df_display_mov.rename(columns={
            'Itens_ENTRADA': 'Itens Comprados (Estoque)',
            'Itens_SAÍDA': 'Itens Vendidos',
            'Custo_ENTRADA': 'Custo de Compra (R$)',
            'Valor_ENTRADA': 'Receita de Compra (R$)', # Nome de exibição para Valor_ENTRADA
            'Valor_SAÍDA': 'Receita de Venda (R$)',
            'Custo_SAÍDA': 'Custo de Venda (R$)'
        }, inplace=True)
                
        # Calcula o Lucro Bruto na Tabela
        df_display_mov['Lucro Bruto (R$)'] = df_pivot['Valor_SAÍDA'] - df_pivot['Custo_SAÍDA']

        # Lista de colunas financeiras a formatar
        cols_financeiras = [
            'Custo de Compra (R$)', 'Receita de Compra (R$)', 'Receita de Venda (R$)', 
            'Custo de Venda (R$)', 'Lucro Bruto (R$)'
        ]

        # CORREÇÃO ROBUSTA para formatação (resolve o TypeError inicial)
        for col in cols_financeiras:
            # 1. Cria uma CÓPIA explícita da Series
            series_to_convert = df_display_mov[col].copy() 
                            
            # 2. Converte a CÓPIA para numérico e preenche NaNs
            converted_series = pd.to_numeric(
                series_to_convert, 
                errors='coerce'
            ).fillna(0)
                            
            # 3. Atribui a Série convertida de volta ao DataFrame usando .loc
            df_display_mov.loc[:, col] = converted_series
                            
            # 4. Aplica a formatação de string (R$ 1.234,56)
            df_display_mov.loc[:, col] = df_display_mov[col].apply(
                lambda x: f"R$ {x:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            )
                            
        st.dataframe(df_display_mov.sort_values(by='Mês/Ano', ascending=False), hide_index=True, use_container_width=True)

    elif report_type == "Análises de Desempenho e Clientes":
        st.subheader("Análises de Desempenho e Clientes")

        df = carregar_dados()

        # Filtrar apenas as vendas concluídas (SAÍDA e PAGO/A RECEBER para receita)
        df_vendas = df[(df['Tipo de Movimentação'] == 'SAÍDA')].copy()

        if df_vendas.empty:
            st.info("Não há dados de vendas suficientes para gerar análises de desempenho.")
        else:
            # --- 1. MELHORES CLIENTES ---
            st.markdown("### 🥇 Top 25 Melhores Clientes (por Receita de Venda)")

            # Agrupar por Cliente e somar a Receita (Preço Venda Total)
            df_top_clientes = df_vendas.groupby('Cliente').agg(
                Receita_Total=('Preço Venda Total', 'sum')
            ).sort_values(by='Receita_Total', ascending=False).reset_index().head(25)
            
            # Formatação
            df_top_clientes['Receita_Total_Display'] = df_top_clientes['Receita_Total'].apply(
                lambda x: f"R$ {x:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            )
            
            st.dataframe(df_top_clientes[['Cliente', 'Receita_Total_Display']].rename(columns={'Receita_Total_Display': 'Receita Total'}), 
                        hide_index=True, use_container_width=True)

            # Gráfico de Clientes (opcional, mas útil)
            fig_clientes = px.bar(
                df_top_clientes, 
                x='Cliente', 
                y='Receita_Total', 
                title='Receita Total por Cliente',
                labels={'Receita_Total': 'Receita Total (R$)'},
                color='Receita_Total',
                color_continuous_scale=px.colors.sequential.Plotly3 # Escala de cor
            )
            fig_clientes.update_traces(hovertemplate='Cliente: %{x}<br>Receita: R$%{y:,.2f}<extra></extra>')
            st.plotly_chart(fig_clientes, use_container_width=True)

            st.markdown("---")
            
            # --- 2. TOP 25 PRODUTOS VENDIDOS ---
            st.markdown("### 📦 Top 25 Produtos Mais Vendidos (por Quantidade)")
            
            # Agrupar por Produto e somar a Quantidade
            df_top_produtos = df_vendas.groupby('Produto').agg(
                Quantidade_Vendida=('Quantidade', 'sum')
            ).sort_values(by='Quantidade_Vendida', ascending=False).reset_index().head(25)
            
            st.dataframe(df_top_produtos, hide_index=True, use_container_width=True)
            
            st.markdown("---")

            # --- 3. MELHORES DIAS DA SEMANA ---
            st.markdown("### 📅 Vendas por Dia da Semana")
            
            # Garantir que 'Data' seja datetime
            df_vendas['Data'] = pd.to_datetime(df_vendas['Data'])
            
            # Criar coluna com o nome do dia da semana em Português
            dias_semana_map = {
                0: 'Segunda-feira', 1: 'Terça-feira', 2: 'Quarta-feira', 
                3: 'Quinta-feira', 4: 'Sexta-feira', 5: 'Sábado', 6: 'Domingo'
            }
            df_vendas['Dia_Semana'] = df_vendas['Data'].dt.dayofweek.map(dias_semana_map)
            
            # Agrupar por Dia da Semana
            df_dias = df_vendas.groupby(['Dia_Semana', df_vendas['Data'].dt.dayofweek]).agg(
                Receita_Media=('Preço Venda Total', 'mean'),
                Total_Transacoes=('ID_Venda', 'count')
            ).reset_index()
            
            # Reordenar corretamente pelo número do dia da semana (coluna oculta)
            df_dias.sort_values(by='Data', inplace=True)
            
            # Formatação para exibição
            df_dias['Receita_Media_Display'] = df_dias['Receita_Media'].apply(
                lambda x: f"R$ {x:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            )

            st.dataframe(
                df_dias[['Dia_Semana', 'Receita_Media_Display', 'Total_Transacoes']].rename(
                    columns={'Dia_Semana': 'Dia da Semana', 'Receita_Media_Display': 'Ticket Médio (R$)', 'Total_Transacoes': 'Total de Vendas'}
                ), 
                hide_index=True, 
                use_container_width=True
            )
            
            # Gráfico de Receita Média por Dia
            fig_dias = px.bar(
                df_dias, 
                x='Dia_Semana', 
                y='Receita_Media', 
                title='Ticket Médio por Dia da Semana',
                labels={'Dia_Semana': 'Dia da Semana', 'Receita_Media': 'Ticket Médio (R$)'},
                category_orders={"Dia_Semana": list(dias_semana_map.values())} # Garante a ordem correta
            )
            fig_dias.update_traces(hovertemplate='Dia: %{x}<br>Ticket Médio: R$%{y:,.2f}<extra></extra>')
            st.plotly_chart(fig_dias, use_container_width=True)
                    
def page_config():
# ... (Restante do código mantido)
    st.markdown("<div class='big-title'>CONFIGURAÇÕES</div>", unsafe_allow_html=True)
    st.markdown("<div class='subtitle'>Informações de Arquivo e Sistema.</div>", unsafe_allow_html=True)
    st.write("Caminhos de arquivos usados pelo app (Estrutura `main.py`):")
    st.write("- Banco de Dados (Excel Único):", ARQUIVO_EXCEL)
    st.markdown("**Dica:** Faça backup manual do arquivo antes de alterações em massa.")


# ===== FUNÇÃO PRINCIPAL (MAIN) DO main2 copy 2.py (Adaptada) =====

def main():
# ... (Restante do código mantido)
    st.set_page_config(page_title="Perfumes & Variedades", layout="wide", initial_sidebar_state="expanded")
        
    inject_css()
    
    st.sidebar.markdown("<div style='display:flex;align-items:center;gap:12px;padding-left:8px;padding-bottom:8px'><div style='font-size:28px;line-height:1'>🧴</div><div><b style='color:#880e4f;font-size:18px'>Perfumes & Variedades</b><div style='font-size:12px;color:#e91e63'>Painel de vendas</div></div></div>", unsafe_allow_html=True)

    
    # Rádio para Menu Lateral - CORRIGIDO O AVISO DE ACESSIBILIDADE
    menu = st.sidebar.radio(
        "Navegação Principal", 
        ["Dashboard", "Vendas","Produtos", "Clientes", "Relatórios", "Configurações"], 
        index=0, # Inicia no Dashboard
        format_func=lambda x: x, 
        label_visibility="hidden" 
    )
    
    if menu == "Dashboard":
        page_dashboard()
    elif menu == "Vendas":
        page_sales()
    elif menu == "Produtos":
        page_products()
    elif menu == "Clientes":
        page_clients()
    elif menu == "Relatórios":
        page_reports()
    elif menu == "Configurações":
        page_config()
    

if __name__ == '__main__':
    main()